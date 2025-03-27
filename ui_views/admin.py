from ui.py.admin_dashboard import Ui_MainWindow
from PyQt6.QtWidgets import QApplication,QMainWindow,QPushButton,QTableWidgetItem,QMessageBox,QHeaderView,QFileDialog
from PyQt6.QtCore import Qt,QSettings
from sqlalchemy.sql import text
from .project import NewProject
from .subtask_dialog import NewSubtask
from .assignment import AssignUser
from db.db import GetCursor
from db.db_pool import DatabasePool
import mysql.connector
# from .extract_log_user_admin import ExtractUserLogAdmin
from openpyxl.styles import Alignment
from db.database_worker import DatabaseWorker
from .edit import EditProject
from .edit_sub import EditSubtask
from .usermanager import UserManager
from .timesheet import Timesheet
from .create_user import CreateUser
from .edit_assignment import EditAssignment
import pandas as pd
from openpyxl.styles import Font
import sys,os
from .logged_in_users import LoggedInUsers
from openpyxl.styles import PatternFill
from .reset_user_password import Reset_User_Password
from openpyxl import Workbook,load_workbook
from datetime import datetime
import datetime as dt
from PyQt6.QtGui import QIcon,QPixmap,QCursor
from .excel_report import Extract_to_Excel
class MyAdmin(QMainWindow,Ui_MainWindow):
    def __init__(self,username):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("Sidebar Menu")
        self.project_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.subtask_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.assignment_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.project_btn_1.clicked.connect(self.switch_to_projectPage)
        self.project_btn_2.clicked.connect(self.switch_to_projectPage)
        self.project_btn_1.setChecked(True)
        self.project_btn_2.setChecked(True)
        self.regularisation_btn_1.clicked.connect(self.switch_to_regularisation_page)
        self.regularisation_btn_2.clicked.connect(self.switch_to_regularisation_page)
        self.label_3.setText(username)
        self.reset_user_password_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        # self.extract_btn.clicked.connect(self.extract_to_excel)
        self.extract_btn.clicked.connect(self.extract_to_excel_2)
        self.assignment_btn_1.clicked.connect(self.switch_to_assignmentPage)
        self.assignment_btn_2.clicked.connect(self.switch_to_assignmentPage)
        self.timesheet_btn_1.clicked.connect(self.switch_to_timesheetPage)
        self.timesheet_btn_2.clicked.connect(self.switch_to_timesheetPage)
        self.setWindowTitle("Admin Dashboard")
        self.add_subtask_dialog=NewSubtask(self.get_project_names())
        self.open_user_assignment_form=AssignUser(self.get_project_names(),self.label_3.text())
        self.open_edit_assignment_form=EditAssignment(self.get_project_names(),self.get_usernames())
       
        # icon_path_on=self.resource_path("assets/icons/messages_white.png")
        # print(icon_path_on)
        # icon_path_off=self.resource_path("assets/icons/messages.png")
        # icon_1=QIcon()
        # icon_1.addPixmap(QPixmap(icon_path_off),QIcon.Mode.Normal, QIcon.State.Off)
        # icon_1.addPixmap(QPixmap(icon_path_on),QIcon.Mode.Normal, QIcon.State.On)
        # self.timesheet_btn_1.setIcon(icon_1)
        self.icon_name_widget.setHidden(True)
        
        self.new_project_btn.clicked.connect(self.open_project_form)
        self.project_edit_btn.clicked.connect(self.open_project_edit_form)
        self.subtask_edit_btn.clicked.connect(self.open_subtask_edit_form)
        self.edit_assign_btn.clicked.connect(self.open_assignment_edit_form)
        self.week_combo.clear()
        self.week_combo.addItems([str(i) for i in range(1,54)])
        # self.populate_timesheet_table()
        self.populate_timesheet_table_3()
        self.week_combo.currentIndexChanged.connect(self.populate_timesheet_table_3)
        self.refresh_btn.clicked.connect(self.populate_timesheet_table_3)
        self.sign_out_btn_1.clicked.connect(self.show_login_window)
        self.sign_out_btn_2.clicked.connect(self.show_login_window)
        self.search_line_edit.textChanged.connect(self.filter_table_by_widget)
        self.threads=[]

        self.add_subtask_btn.clicked.connect(self.add_subtask_form)
        self.assign_project_btn.clicked.connect(self.user_assignment_form)
        self.project_delete_btn.clicked.connect(self.delete_selected_project_row)
        self.subtask_delete_btn.clicked.connect(self.delete_selected_subtask_row)
        self.delete_assign_btn.clicked.connect(self.delete_selected_assignment_row)
        self.create_user_btn.clicked.connect(self.open_create_user_form)
        self.reset_user_password_btn.clicked.connect(self.open_reset_password_dialog)
        self.logged_in_btn.clicked.connect(self.show_logged_in_users)
        self.stackedWidget.setCurrentIndex(0)

        self.fetch_data()
        self.fetch_subtask_data()
        self.fetch_assigned_projects()
        self.load_open_requests()
        self.timesheet_table.itemDoubleClicked.connect(self.show_timesheet)
        self.accept_btn.clicked.connect(self.accept_regularisation_request)

    # def resource_path(relative_path):
    #     if hasattr(sys,'_MEIPASS'):
    #         return os.path.join(sys._MEIPASS,relative_path)
    #     return os.path.join(os.path.abspath("."),relative_path)
    # def resource_path(self,relative_path):
    #     base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    #     base_path = os.path.abspath(os.path.join(base_path, ".."))
    #     return os.path.join(base_path, relative_path)


    def open_reset_password_dialog(self):
        repoting_to=self.label_3.text()
        # dialog=Reset_User_Password()
        # dialog.exec()
        query = text("SELECT username FROM password_reset WHERE reporting_to = :reporting_to")
        params = {"reporting_to": repoting_to}
        worker = DatabaseWorker((query, params))
        worker.result_ready.connect(self.handle_reset_password_data)
        worker.error_occured.connect(self.handle_error)
        self.threads.append(worker)
        worker.finished.connect(lambda:self.threads.remove(worker))
        worker.start()


    def load_open_requests(self):
        db_instance=DatabasePool(pool_type="read")
        
        with db_instance.get_db_connection() as conn:
            cursor=conn.cursor()
            try:
                query="select employee_name,date,reason from regularisation"
                cursor.execute(query)
                result=cursor.fetchall()
                if not result:
                    return
                self.regularisation_table.setRowCount(len(result))
                for row_idx,row_data in enumerate(result):
                    for col_idx,cell_data in enumerate(row_data):
                        item=QTableWidgetItem(str(cell_data) if cell_data else "-")
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                        self.regularisation_table.setItem(row_idx,col_idx,item)
                self.regularisation_table.resizeColumnsToContents()
                self.regularisation_table.horizontalHeader().setStretchLastSection(True)
                
            except mysql.connector.Error as err:
                
                print(err)
    def accept_regularisation_request(self):
        selected_row=self.regularisation_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self,"No Selection","Please select a request from the table.")
            return
        
        employee_name=self.regularisation_table.item(selected_row,0).text()
        regularisation_date=self.regularisation_table.item(selected_row,1).text()

        reply=QMessageBox.question(self,"Confirm Regularisation",f"Do you want to accept the regularisation for {employee_name} on {regularisation_date}?",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
        if reply==QMessageBox.StandardButton.Yes:
            db_instance=DatabasePool(pool_type="write")
            with db_instance.get_db_connection() as conn:
                cursor=conn.cursor()
                try:
                    #Fetch login time from user_logs
                    cursor.execute("""select login_time from user_logs where employee_name = %s and login_date = %s   """,(employee_name,regularisation_date))
                    result=cursor.fetchone()
                    if not result:
                        QMessageBox.critical(self,"Error","Login time not found.")
                        return
                    login_time=result[0]

                    #Convert login time and calculate hours worked
                    login_time_dt=datetime.strptime(str(login_time),"%H:%M:%S")
                    logout_time_dt=datetime.strptime("18:00:00","%H:%M:%S")

                    #Compute hours worked
                    if login_time_dt>logout_time_dt:
                        total_hours=8
                    else:
                        worked_time=logout_time_dt - login_time_dt
                        total_hours=worked_time.total_seconds()/3600 #convert to hours
                    #Update user_logs table
                    update_query="""
update user_logs set state='closed',regularised='yes', logout_time='18:00:00',hours_logged=%s where employee_name=%s and login_date=%s

"""
                    cursor.execute(update_query,(total_hours,employee_name,regularisation_date))
                    #Delete from regularisation table
                    delete_query="""
Delete from regularisation where employee_name =%s and date=%s



"""
                    cursor.execute(delete_query,(employee_name,regularisation_date))
                    conn.commit()
                    QMessageBox.information(self,"Success","Regularisation request accepted successfully!")
                    self.regularisation_table.removeRow(selected_row)
                except mysql.connector.Error as err:
                    conn.rollback()
                    QMessageBox.critical(self,"Error","Failed to accept regularisation request.")
                    print("Database error: ",err)



    def show_logged_in_users(self):
        self.logged_in_users=LoggedInUsers()
        self.logged_in_users.exec()

    def handle_reset_password_data(self,result):
        if not result:
            dialog=Reset_User_Password([])
            dialog.form_data_submitted.connect(self.save_new_password)
            dialog.exec()
            return
        usernames=[row[0] for row in result]
        dialog=Reset_User_Password(usernames)
        dialog.form_data_submitted.connect(self.save_new_password)
        dialog.exec()

    def save_new_password(self,obj,employee,newpass,confirmpass):
        if employee.lower()=="--select user--":
            QMessageBox.warning(self,"No Selection","Please select a user to reset password!")
            return
        
        if newpass != confirmpass:  # Case-sensitive check
            QMessageBox.warning(self, "Password Mismatch", "Passwords do not match!")
            return
        update_query = text("UPDATE users SET password = :newpass WHERE BINARY username = :employee")
        params = {"newpass": newpass, "employee": employee}
        worker = DatabaseWorker((update_query, params))
        worker.set_write_operation(True)
        worker.result_ready.connect(lambda result: self.handle_password_update(result, employee,obj))
        worker.error_occured.connect(self.handle_error)
        self.threads.append(worker)
        worker.finished.connect(lambda:self.threads.remove(worker))
        worker.start()

    def handle_password_update(self,result,employee,obj):
        if result =="write operation successfull.":
             delete_query = text("DELETE FROM password_reset WHERE BINARY username = :employee")
             params = {"employee": employee}
             worker = DatabaseWorker((delete_query, params))
             worker.set_write_operation(True)
             worker.result_ready.connect(lambda result: self.close_reset_password_dialog(result,obj,employee))
             worker.error_occured.connect(self.handle_error)
             self.threads.append(worker)
             worker.finished.connect(lambda:self.threads.remove(worker))
             worker.start()
        else:
            QMessageBox.critical(self,"Update Failed","Failed to update the password. User might not exist.")


    def close_reset_password_dialog(self,result,obj,employee):
        QMessageBox.information(self, "Success", f"Password reset for {employee} was successful!")
        obj.close()

        















    def extract_to_excel_2(self):
        reporting_to=self.label_3.text()
        dialog=Extract_to_Excel(reporting_to)
        dialog.form_data_submitted.connect(self.extract_employee_data)
        dialog.exec()
        
    def extract_employee_data(self,filter_by,start_mon_week,end_mon_week,employee):
        query="Select week,username,project, pspelement,segment,subtask,logged_hours from tasks where"
        params={}
        month_to_week = {
        "january":  (1, 5), "february":  (5, 9), "march":  (9, 14), "april":  (14, 18),
        "may":  (18, 22), "june":  (22, 27), "july":  (27, 31), "august":  (31, 35),
        "september":  (36, 40), "october":  (40, 44), "november":  (44, 48), "december":  (49, 53)
    }
        if filter_by.lower()=="week":
            query += " week between :start_week and :end_week"
            params["start_week"]=start_mon_week
            params["end_week"]=end_mon_week

        elif filter_by.lower()=="month":
            start_month=start_mon_week.lower()
            end_month=end_mon_week.lower()
            if start_month in month_to_week and end_month in month_to_week:
                start_week=month_to_week[start_month][0]
                end_week=month_to_week[end_month][1]
                query +=" week between :start_week and :end_week"
                params["start_week"]=start_week
                params["end_week"]=end_week
            else:
                QMessageBox.warning(self,"Invalid Month","One or both selected months are invalid!")
                return
        elif filter_by.lower()=="--extract by month or week--":
            QMessageBox.warning(self,"No Selection","Please select month or week.")
            return
        if  employee.lower()!="all employees":
            query+=" and username=:employee"
            params["employee"]=employee
        worker=DatabaseWorker((text(query),params))
        worker.result_ready.connect(self.handle_employee_data)
        worker.error_occured.connect(self.handle_error)
        worker.finished.connect(lambda:self.threads.remove(worker))
        self.threads.append(worker)
        worker.start()

    def handle_employee_data(self,data):
        if not data:
            QMessageBox.warning(self,"No Data","There is no data to export.")
            return
        headers=["Week","Employee Name","Project","PSP","Segment","Subtask","Logged Hours"]
        df=pd.DataFrame(data,columns=headers)
        file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_path:
            try:
                df.to_excel(file_path, index=False, engine="openpyxl")  # Save as Excel
                 # Load workbook and select active sheet
                wb = load_workbook(file_path)
                ws = wb.active

            # Set column widths
                column_widths = [10, 20, 25, 15, 20, 25, 15]  # Adjust as needed
                for col_num, width in enumerate(column_widths, start=1):
                    ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

            # Apply background color to headers
                header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                for cell in ws[1]:  # First row (headers)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")  # Center headers

            # Center-align all data
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            # Save workbook
                wb.save(file_path)

                QMessageBox.information(self, "Success", f"Data successfully exported to:\n{file_path}")
        
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save file:\n{str(e)}")
            #     QMessageBox.information(self, "Success", f"Data successfully exported to:\n{file_path}")
            # except Exception as e:
            #     QMessageBox.critical(self, "Error", f"Failed to save file:\n{str(e)}")
        
    def filter_table_by_widget(self):
        idx=int(self.stackedWidget.currentIndex())
        if idx==0:
            self.filter_table(self.project_table)
            self.filter_table(self.subtask_table)
        elif idx==1:
            self.filter_table(self.assignment_table)
        elif idx==2:
            self.filter_table(self.timesheet_table)
        elif idx==3:
            self.filter_table(self.regularisation_table)
        else:
            pass


    def filter_table(self,tablename):
        # print(self.stackedWidget.currentIndex())
        search_text=self.search_line_edit.text().lower()
        for row in range(tablename.rowCount()):
            show_row=False
            for col in range(tablename.columnCount()):
                item=tablename.item(row,col)
                if item and search_text in item.text().lower():
                    show_row=True
                    break
            tablename.setRowHidden(row,not show_row)

    def extract_to_excel(self):
        week_number=self.week_combo.currentText()
        workbook=Workbook()
        sheet=workbook.active
        sheet.title="TimeSheet"
        sheet.column_dimensions['A'].width=20
        sheet.column_dimensions['B'].width=20
        sheet.column_dimensions['C'].width=20
        sheet.column_dimensions['D'].width=20
        bold_font=Font(bold=True)
        fill=PatternFill(start_color="73a4ca",end_color="73a4ca",fill_type="solid")


   
        # data=[]
        # for row in range(rows):
        #     row_data=[]
        #     for col in range(cols):
        #         item=self.timesheet_table.item(row,col)
        #         row_data.append(item.text() if item else "")
        #     data.append(row_data)

        # df=pd.DataFrame(data,columns=[self.timesheet_table.horizontalHeaderItem(col).text() for col in range(cols)])
        # df["week"]=week_number
        # file_path,_=QFileDialog.getSaveFileName(self,"Save File","","Excel Files (*.xlsx)")
        # if file_path:
        #     with pd.ExcelWriter(file_path,engine="openpyxl") as writer:
        #         df.to_excel(writer,index=False,sheet_name=f"Week_{week_number}")
        #     print(f"Data exported successfully to {file_path}")
        sheet.cell(row=1,column=1).value="Week"
        sheet.cell(row=1,column=2).value=week_number
        sheet.cell(row=1,column=1).font=bold_font
        sheet.cell(row=1,column=2).font=bold_font
        for column in range(self.timesheet_table.columnCount()):
            sheet.cell(row=2,column=column+1).value=self.timesheet_table.horizontalHeaderItem(column).text()
            sheet.cell(row=2,column=column+1).font=bold_font
            sheet.cell(row=2,column=column+1).fill=fill
            sheet.cell(row=2,column=column+1).alignment=Alignment(horizontal="center")
        
        for row in range(self.timesheet_table.rowCount()):
            for col in range(self.timesheet_table.columnCount()):
                item=self.timesheet_table.item(row,col)
                sheet.cell(row=row+3,column=col+1).value=item.text() if item else ""
        
        options=QFileDialog.Option.DontUseNativeDialog
        file_path,_=QFileDialog.getSaveFileName(None,"Save Excel File","","Excel Files(*.xlsx)",options=options)
        if file_path:
            workbook.save(file_path)
            # print("success")

    
    def show_login_window(self):
        from .login import Login
        confirmation=QMessageBox.question(self,"Confirm sign out",f"Are you sure you want to sign out?",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
        if confirmation !=  QMessageBox.StandardButton.Yes:
            return
        # Capture logout time
        # logout_time=dt.datetime.now().strftime("%H:%M:%S")
        # print("The log out time is ",logout_time)

        # #Remove login token from the database
        # db_instance=DatabasePool(pool_type="write")
        # with db_instance.get_db_connection() as conn:
        #     cursor=conn.cursor()
        #     try:
        #         query="update users set login_token=NULL, token_expiry=NULL where username=%s"
        #         cursor.execute(query,(self.label_3.text(),))
        #         print(f"The user is -{self.label_3.text()}-")

        #         query="select login_time from user_logs where employee_name=%s order by login_time desc limit 1"
        #         cursor.execute(query,(self.label_3.text(),))
        #         login_hour=cursor.fetchone()[0]
        #         print("the login hour is ",login_hour)
        #         total_logged_hours=dt.datetime.strptime(logout_time,"%H:%M:%S")-login_hour
        #         query="update user_logs set logout_time=%s,hours_logged=%s where employee_name=%s and login_time=%s"
        #         cursor.execute(query,(logout_time,total_logged_hours,self.label_3.text(),login_hour))

        #         conn.commit()
        #     except mysql.connector.Error as err:
        #         conn.rollback()
        #         QMessageBox.critical(self,"Database Error",f"An error occurred: {err}")
        
        settings=QSettings("Schueco","app")
        settings.remove("login_token")


        login_dialog=Login()
        self.close()
        login_dialog.exec()
        # pass

    def show_timesheet(self,item):
        row=item.row()
        week=self.week_combo.currentText()
        user=self.timesheet_table.item(row,2).text()
        timesheet=Timesheet(user,week)
        # timesheet.cal_da
        timesheet.exec()

    def delete_selected_assignment_row(self):
        selected_row=self.assignment_table.currentRow()
        assigned_by=self.label_3.text()
        if selected_row==-1:
            QMessageBox.warning(self,"No Selection","Please select a row to delete.")
            return
        assignment_id=self.assignment_table.item(selected_row,0).text()
        employee_id=self.assignment_table.item(selected_row,1).text()
        project_name=self.assignment_table.item(selected_row,3).text()
        subtask=self.assignment_table.item(selected_row,4).text()
        confirmation=QMessageBox.question(self,"Confirm deletion",f"Are you sure you want to delete assignment {assignment_id}? ",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
        if confirmation != QMessageBox.StandardButton.Yes:
            return
        # try:
        #     conn=mysql.connector.connect(
        #     host="10.95.136.128",
        #     user="Meeraj", 
        #     password="Mohammed&meeraj786", 
        #     database="timesheet",
        #     port=3306
                

        #     )
        #     cursor=conn.cursor()
        #     query="delete from assignments where user_id=%s and project=%s and subtask=%s and assigned_by=%s"
        #     values=(employee_id,project_name,subtask,assigned_by)
        #     cursor.execute(query,values)
        #     conn.commit()
        #     if cursor.rowcount>0:
        #         self.assignment_table.removeRow(selected_row)
        #         QMessageBox.information(self,"Success","Deletion Successful.")
        #     else:
        #         QMessageBox.critical(self,"Failed","Deletion failed.")
        # except mysql.connector.Error as err:
        #     QMessageBox.critical(self,"Failed",f"{err.msg}")
        # finally:
        #     if conn.is_connected():
        #         cursor.close()
        #         conn.close()
        query=text("delete from assignments where user_id=:user_id and project=:project and subtask=:subtask and assigned_by=:assigned_by")
        params={"user_id":employee_id,
                "project":project_name,
                "subtask":subtask,
                "assigned_by":assigned_by
                }
        worker=DatabaseWorker((query,params))
        def handle_deletion_success(message):
            QMessageBox.information(self,"Success","Deletion Successful.")
            selected_row=self.assignment_table.currentRow()
            self.assignment_table.removeRow(selected_row)
            
        worker.set_write_operation(True)
        worker.result_ready.connect(handle_deletion_success)
        worker.error_occured.connect(self.handle_error)
        worker.finished.connect(lambda:self.threads.remove(worker))
        self.threads.append(worker)
        worker.start()
        
        
    
    def open_create_user_form(self):
        self.create_user_dialog=CreateUser()
        self.create_user_dialog.form_data_submitted.connect(self.create_user)
        self.create_user_dialog.exec()
    

    def create_user(self,create_user_ob,username,employee_id,email,password,repeat_password,role,reporting_to):
        # success,message=self.handle_user_creation(username,email,password,repeat_password,role)
        user_manager=UserManager(self)
        is_valid,message=user_manager.validate_data(create_user_ob,username,employee_id,email,password,repeat_password,role,reporting_to)
        if not is_valid:
            QMessageBox.warning(self,"Validation Error",message)
            return
        user_manager.check_user_existence(create_user_ob,username,employee_id,email,password,role,reporting_to)
        
    
    # def handle_user_creation(self,username,email,password,repeat_password,role):
    #     is_valid,message=self.validate_data(username,email,password,repeat_password,role)
    #     if not is_valid:
    #         return False,message
        

    
    # def validate_data(self,username,email,password,repeat_password,role):
    #     if self.user_exists(username,email):
    #         return False,"User already exists!"
        

    # def user_exists(self,username,email):
    #     query=text("select count(*) from users where username=:username or email=:email")
    #     self.worker=DatabaseWorker(query=(query,{'username':username,'email':email}))
    #     data=self.worker.result_ready.connect(self.check_user_exists)
    #     return data[0]>1

    # def check_user_exists(self,data):
    #     return data
    #-------------------------------------------------------------------------------------------------
    def open_assignment_edit_form(self):
        # self.assignment_edit_dialog=EditAssignment(self.get_project_names,self.get_usernames)
        self.open_edit_assignment_form=EditAssignment(self.get_project_names(),self.get_usernames())
        try:
            username,project,subtask,start_date,end_date,duration=self.retrive_assignment_details()
            formated_start_date=datetime.strptime(start_date,"%Y-%m-%d").strftime("%d-%m-%Y")
            formated_end_date=datetime.strptime(end_date,"%Y-%m-%d").strftime("%d-%m-%Y")
        except TypeError:
            return
        self.open_edit_assignment_form.ui.edit_user_combo.setCurrentText(username)
        self.open_edit_assignment_form.ui.edit_project_combo.setCurrentText(project)
        self.open_edit_assignment_form.ui.edit_subtask_combo.setCurrentText(subtask)
        self.open_edit_assignment_form.ui.edit_start_date_le.setText(formated_start_date)
        self.open_edit_assignment_form.ui.edit_end_date_le.setText(formated_end_date)
        self.open_edit_assignment_form.ui.edit_duration_le.setText(duration)
        self.open_edit_assignment_form.form_data_submitted.connect(self.edit_assignmets)
        self.open_edit_assignment_form.exec()

    

    # def populate_timesheet_table(self):
    #     week=self.week_combo.currentText()
    #     reporting_to=self.label_3.text()
    #     db=GetCursor()
    #     if db.conn and db.cursor:
    #         query="select username,user_id,total_hours from weekly_hours where week=%s and reporting_to=%s"
    #         db.cursor.execute(query,(week,reporting_to))
    #         results=db.cursor.fetchall()
    #         self.populate_timsheet_table_2(results)
    #         print("yess")
    def populate_timesheet_table_3(self):
        week=self.week_combo.currentText()
        reporting_to=self.label_3.text()
        query=text("select week,username,project,pspelement,segment,subtask,logged_hours from tasks where week=:week and reporting_to=:reporting_to")
        self.worker=DatabaseWorker(query=(query,{'week':week,'reporting_to':reporting_to}))
        self.worker.result_ready.connect(self.populate_timsheet_table_2)
        self.worker.error_occured.connect(self.handle_error)
        self.worker.start()
        # query=text("select username,user_id,total_hours from weekly_hours where week=:week and reporting_to=:reporting_to")
        # self.worker=DatabaseWorker(query=(query,{'week':week,'reporting_to':reporting_to}))
        # self.worker.result_ready.connect(self.populate_timsheet_table_2)
        # self.worker.error_occured.connect(self.handle_error)
        # self.worker.start()

    
    def handle_error(self,error):
        QMessageBox.warning(self,"Error",f"{error}")


    def populate_timsheet_table_2(self,data):
        # self.timesheet_table.setRowCount(len(data))    
        # for row_idx, row_data in enumerate(data):
        #     for col_idx, value in enumerate(row_data,1):
        #         item=QTableWidgetItem(str(value))
        #         item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        #         self.timesheet_table.setItem(row_idx,col_idx,item)
        # self.timesheet_table.resizeColumnsToContents()
        # self.update_serial_number(self.timesheet_table)
        db_to_widget_mapping={
            "week":"Week",
            "username":"Employee Name",
            "project":"Project",
            "pspelement":"PSP",
            "segment":"Segment",
            "subtask":"Subtask",
            "logged_hours":"Logged Hours"
            
        }
        widget_headers = [self.timesheet_table.horizontalHeaderItem(col).text() for col in range(self.timesheet_table.columnCount())]
        header_to_data_index = {widget_header: idx for idx, (db_col, widget_header) in enumerate(db_to_widget_mapping.items()) if widget_header in widget_headers}
        self.timesheet_table.setRowCount(len(data))
        for row_idx, row_data in enumerate(data):
            for col_idx, header in enumerate(widget_headers):
                if header in header_to_data_index:
                    data_index = header_to_data_index[header]  # Get the index of the MySQL column in the data
                    value = row_data[data_index]  # Fetch the value using the index
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.timesheet_table.setItem(row_idx, col_idx, item)
        # self.timesheet_table.resizeColumnsToContents()
        header = self.timesheet_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.update_logged_hours()
        self.update_serial_number(self.timesheet_table)
        


    def update_logged_hours(self):
        pass

    def open_project_edit_form(self):
        self.edit_dialog=EditProject()
        try:
            project_name,psp_element=self.retrieve_project_psp_name()
        except TypeError:
            return
        self.edit_dialog.ui.edit_project_name_le.setText(project_name)
        self.edit_dialog.ui.edit_psp_element_le.setText(psp_element)
        self.edit_dialog.form_data_submitted.connect(self.edit_project)
        self.edit_dialog.exec()
    def edit_assignmets(self,obj,username,project,subtask,start_date,end_date,duration):
        selected_row=self.assignment_table.currentRow()
        assignment_id=self.assignment_table.item(selected_row,0).text()
        assigned_by=self.label_3.text()
        # try:
            
        #     conn=mysql.connector.connect(
        #         host="10.95.136.128",
        #         user="Meeraj",
        #         password="Mohammed&meeraj786",
        #         database="timesheet"

        #     )
        #     cursor=conn.cursor()
        #     query="update assignments set username=%s,project=%s,subtask=%s,start_date=STR_TO_DATE(%s, '%d-%m-%Y'),end_date=STR_TO_DATE(%s, '%d-%m-%Y'),duration=%s where assignment_id=%s "
        #     values=(username,project,subtask,start_date,end_date,duration,assignment_id)
        #     cursor.execute(query,values)
        #     conn.commit()
        #     query="Select assignment_id,user_id,username,project,subtask,start_date,end_date,duration from assignments where assigned_by=%s"
        #     cursor.execute(query,(assigned_by,))
        #     data=cursor.fetchall()
        #     self.populate_table(data,self.assignment_table)

        # except mysql.connector.Error as err:
        #     QMessageBox.warning(self,"Failed",f"{err.msg}")
        # finally:
        #     if conn.is_connected():
        #         cursor.close()
        #         conn.close()
        query = text("""
        UPDATE assignments
        SET username = :username, project = :project, subtask = :subtask,
            start_date = STR_TO_DATE(:start_date, '%d-%m-%Y'),
            end_date = STR_TO_DATE(:end_date, '%d-%m-%Y'),
            duration = :duration
        WHERE assignment_id = :assignment_id
    """)
        params = {
        'username': username,
        'project': project,
        'subtask': subtask,
        'start_date': start_date,
        'end_date': end_date,
        'duration': duration,
        'assignment_id': assignment_id
    }
        def handle_edit_success(message):
            obj.close()
            assigned_by = self.label_3.text()
            query = text("""
        SELECT assignment_id, user_id, username, project, subtask, start_date, end_date, duration
        FROM assignments
        WHERE assigned_by = :assigned_by
    """)
            params = {'assigned_by': assigned_by}
            worker=DatabaseWorker((query,params))
            def populate_table_after_edit(data):
                self.populate_table(data,self.assignment_table)
                
            worker.result_ready.connect(populate_table_after_edit)
            worker.error_occured.connect(self.handle_error)
            worker.finished.connect(lambda:self.threads.remove(worker))
            self.threads.append(worker)
            worker.start()

        db_worker = DatabaseWorker((query,params))
        db_worker.set_write_operation(True)

        db_worker.result_ready.connect(handle_edit_success)
        db_worker.error_occured.connect(self.handle_error)
        db_worker.finished.connect(lambda:self.threads.remove(db_worker))
        self.threads.append(db_worker)
        db_worker.start()



    def open_subtask_edit_form(self):
        self.edit_subtask_dialog=EditSubtask(self.get_project_names())
        try:
            project_name,psp,subtask=self.retrieve_subtask_data()
            # print(project_name)
            # print(psp)
            # print(subtask)
        except TypeError:
            return
        self.edit_subtask_dialog.ui.project_combo_box_edit.setCurrentText(project_name)
        self.edit_subtask_dialog.ui.psp_element_le_edit.setText(psp)
        self.edit_subtask_dialog.ui.subtask_name_le_edit.setText(subtask)
        self.edit_subtask_dialog.form_data_submitted.connect(self.edit_and_update_subtask)
        self.edit_subtask_dialog.exec()

    def edit_and_update_subtask(self,project_name,subtask,psp):
        # print("data recieverd")
        # print(project_name,psp,subtask)
        self.update_subtask_table(project_name,subtask,psp)
    def update_subtask_table(self,project_name,psp,subtask):
        created_by=self.label_3.text()
        # print("project",project_name)
        # print("subtask",subtask)
        # print("psp",psp)
        selected_row=self.subtask_table.currentRow()
        subtask1=self.subtask_table.item(selected_row,3).text()
        project=self.subtask_table.item(selected_row,1).text()
        try:
            conn=mysql.connector.connect(
                # host="10.95.136.128",
                # user="Meeraj", 
                # password="Mohammed&meeraj786", 
                # database="timesheet",
                # port=3306,
                host="10.95.136.128",
                user="app_user", 
                password="Mohammed&meeraj786", 
                database="timesheet",
               

            )
            cursor=conn.cursor()
            query="select id from projects where project_name=%s and created_by=%s"
            cursor.execute(query,(project_name,created_by))
            i=cursor.fetchone()
            id=i[0]


            query="update subtasks set project_id=%s,project_name=%s,pspelement=%s,subtask=%s where project_name=%s and subtask=%s and created_by=%s"
            values=(id,project_name,psp,subtask,project,subtask1,created_by)
            cursor.execute(query,values)
            conn.commit()
            # print("Subtasks updated successfully")
            self.fetch_subtask_data()

        except mysql.connector.Error as err:
            print("Error occured",err)
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
    def retrieve_project_psp_name(self):
        selected_row=self.project_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self,"No Selection","Please select a row to edit.")
            return
        project_name=self.project_table.item(selected_row,1).text()
        psp_element=self.project_table.item(selected_row,2).text()
        return project_name,psp_element
    
    def retrive_assignment_details(self):
        selected_row=self.assignment_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self,"No Selection","Please select a row to edit.")
            return
        username=self.assignment_table.item(selected_row,2).text()
        project=self.assignment_table.item(selected_row,3).text()
        subtask=self.assignment_table.item(selected_row,4).text()
        start_date=self.assignment_table.item(selected_row,5).text()
        end_date=self.assignment_table.item(selected_row,6).text()
        duration=self.assignment_table.item(selected_row,7).text()
        return username,project,subtask,start_date,end_date,duration

        
    def retrieve_subtask_data(self):
        selected_row=self.subtask_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self,"No Selection","Please select a row to edit.")
            return
        project_name=self.subtask_table.item(selected_row,1).text()
        psp_element=self.subtask_table.item(selected_row,2).text()
        subtask=self.subtask_table.item(selected_row,3).text()
        return project_name,psp_element,subtask
        

    def edit_project(self,data1,data2):
        # print(f"{data1} {data2}")
        self.update_project_table(data1=data1,data2=data2)

    def update_project_table(self,data1,data2):
        selected_row=self.project_table.currentRow()
        project_id=self.project_table.item(selected_row,0).text()
        try:
            conn=mysql.connector.connect(
                host="10.95.136.128",
                user="app_user", 
                password="Mohammed&meeraj786", 
                database="timesheet",

            )
            cursor=conn.cursor()
            query="update projects set project_name=%s,pspelement=%s where id=%s"
            values=(data1,data2,project_id)
            cursor.execute(query,values)
            conn.commit()
            # print("edited successfully")
            self.fetch_data()
            self.populate_project_combo_box()
        except mysql.connector.Error as err:
            print("Error occured")
        except Exception as e:
            print("Unexpected error")
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()

    def get_project_names(self):
        # Retrive all project names from the db
        created_by=self.label_3.text()
        # print("created by is",created_by)
        try:
            conn=mysql.connector.connect(
                host="10.95.136.128",
            user="app_user",
            password="Mohammed&meeraj786",
            database="timesheet"
            )
            cursor=conn.cursor()
            query="select project_name from projects where created_by=%s"
            cursor.execute(query,(created_by,))
            results=cursor.fetchall()
            a=[row[0] for row in results]
            # print("a is ",a)
            return a
        except mysql.connector.Error as err:
            print("Error fetching projects names", err)
            return[]
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
    def get_usernames(self):
        reporting_to=self.label_3.text()
        try:
            conn=mysql.connector.connect(
                host="10.95.136.128",
            user="app_user",
            password="Mohammed&meeraj786",
            database="timesheet"
            )
            cursor=conn.cursor()
            query="select username from users where reporting_to=%s"
            cursor.execute(query,(reporting_to,))
            results=cursor.fetchall()
            return [row[0] for row in results]
        except mysql.connector.Error as err:
            print("Error fetching user names", err)
            return[]
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()

    def get_subtasks(self):
        try:
            conn=mysql.connector.connect(
            host="10.95.136.128",
            user="app_user",
            password="Mohammed&meeraj786",
            database="timesheet"
            )
            cursor=conn.cursor()
            query="select subtask from subtasks"
            cursor.execute(query)
            results=cursor.fetchall()
            return [row[0] for row in results]
        except mysql.connector.Error as err:
            print("Error fetching subtasks", err)
            return[]
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
       
    def populate_project_combo_box(self):
        try:
            conn = mysql.connector.connect(
            host="10.95.136.128",
            user="app_user", 
            password="Mohammed&meeraj786", 
            database="timesheet",
            
            )
            cursor = conn.cursor()
            query = "SELECT project_name FROM projects"
            cursor.execute(query)
            results=cursor.fetchall()
            self.add_subtask_dialog.ui.project_combo_box.clear()
            for row in results:
                print(row[0])
                self.add_subtask_dialog.ui.project_combo_box.addItem(row[0])
        except mysql.connector.Error as e:
            print("Error occured")
    # def closeEvent(self, event):
        
    #     reply=QMessageBox.question(self,"Confirm Exit","Are you sure you want to close the app?",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
    #     if reply==QMessageBox.StandardButton.Yes:
    #         event.accept()
    #     else:
    #         event.ignore()
    def delete_selected_project_row(self):
        selected_row=self.project_table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self,"No Selection","Please select a row to delete.")
            return
        project_id=self.project_table.item(selected_row,0).text()
        project_name=self.project_table.item(selected_row,1).text()
        confirmation=QMessageBox.question(self,"Confirm deletion",f"Are you sure you want to delete project {project_name}?",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
        if confirmation != QMessageBox.StandardButton.Yes:
            return
        try:
            conn = mysql.connector.connect(
            host="10.95.136.128",
            user="app_user", 
            password="Mohammed&meeraj786", 
            database="timesheet",
           
            
        )
            cursor=conn.cursor()
            query="delete from projects where id = %s "
            cursor.execute(query,(project_id,))
            conn.commit()
            if cursor.rowcount>0:
                self.project_table.removeRow(selected_row)
                self.populate_project_combo_box()
                QMessageBox.information(self,"Success","Project deleted successfully!")
            else:
                QMessageBox.warning(self,"Failed","Failed to delete project from the database.")

        except mysql.connector.Error as err:
            QMessageBox.critical(self,"Database Error",f"An error occured: {err}")
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()

    def delete_selected_subtask_row(self):
        selected_row=self.subtask_table.currentRow()
        if selected_row==-1:
            QMessageBox.warning(self,"No Selection","Please select a subtask to delete.")
            return
        subtask=self.subtask_table.item(selected_row,3).text()
        psp_element=self.subtask_table.item(selected_row,2).text()
        confirmation=QMessageBox.question(self,"Confirm deletion",f"Are you sure you want to delete subtask {subtask}?",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
        if confirmation != QMessageBox.StandardButton.Yes:
            return
        try:
            conn = mysql.connector.connect(
            host="10.95.136.128",
            user="app_user", 
            password="Mohammed&meeraj786", 
            database="timesheet",
           
        )
            cursor=conn.cursor()
            query="delete from subtasks where subtask=%s and pspelement=%s and created_by=%s"
            created_by=self.label_3.text()
            cursor.execute(query,(subtask,psp_element,created_by))

            conn.commit()
            if cursor.rowcount>0:
                self.subtask_table.removeRow(selected_row)
                QMessageBox.information(self,"Success","Subtask deleted successfully.")
            else:
                QMessageBox.warning(self,"Failed","Failed to delete subtask from the database.")
        except mysql.connector.Error as err:
            QMessageBox.critical(self,"Failed",f"An error occured:{err}")
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()

    def switch_to_projectPage(self):
        self.stackedWidget.setCurrentIndex(0)
    def switch_to_assignmentPage(self):
        self.stackedWidget.setCurrentIndex(1)
    def switch_to_timesheetPage(self):
        self.stackedWidget.setCurrentIndex(2)
    def switch_to_regularisation_page(self):
        self.stackedWidget.setCurrentIndex(3)
    def open_project_form(self):
        self.dialog=NewProject()
        self.dialog.form_data_submitted.connect(self.add_project)
        self.dialog.exec()
    def add_project(self,data1,data2):
        print("project name ",data1)
        print("psp element ",data2)
        self.insert_into_database(data1=data1,data2=data2)

    def insert_into_database(self,data1,data2):
        admin_name=self.label_3.text()
        
        try: 
            conn = mysql.connector.connect(
                host="10.95.136.128",
                user="app_user", 
                password="Mohammed&meeraj786", 
                database="timesheet",
               
                
            )
            
            cursor = conn.cursor()
            query="insert into projects (project_name,pspelement,created_by) values (%s,%s,%s)"
            values=(data1,data2,admin_name)
            cursor.execute(query,values)
            conn.commit()
            print("Data inserted successfully!")
            self.fetch_data()
            self.populate_project_combo_box()
        except mysql.connector.Error as err:
            QMessageBox.critical(self,"Error occured",f"{err}")
        except Exception as ex:
            print("An unexpected error occurred:", ex)
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
    def fetch_data(self):
        created_by=self.label_3.text()
        # print("printed by ",created_by)
        try:
            conn = mysql.connector.connect(
                host="10.95.136.128",
                user="app_user", 
                password="Mohammed&meeraj786", 
                database="timesheet",
                
            )
            cursor=conn.cursor()
            query="select id,project_name,pspelement,created_date from projects where created_by=%s"
            cursor.execute(query,(created_by,))
            results=cursor.fetchall()
            self.populate_table(results,self.project_table)

        except mysql.connector.Error as err:
            print("Error occured")
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
            
    def populate_table(self,data,table):
        table.setRowCount(len(data))
        for row_idx,row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                item=QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(row_idx,col_idx,item)
        # self.update_serial_number(self.project_table)
        table.resizeColumnsToContents()
    def update_serial_number(self,table_name):
        for row in range(table_name.rowCount()):
            # table_name.setItem(row,0,QTableWidgetItem(str(row+1)))
            item = QTableWidgetItem(str(row + 1))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Align text to center
            table_name.setItem(row, 0, item)
            
        # pass

    def add_subtask_form(self):
        if self.add_subtask_dialog:
            self.add_subtask_dialog.update_project_names(self.get_project_names())
        else:
            self.add_subtask_dialog=NewSubtask(self.get_project_names())
        self.add_subtask_dialog.form_data_submitted.connect(self.add_subtask)
        self.add_subtask_dialog.exec()
    def add_subtask(self,project_name,subtask,psp):
        print(project_name)
        print(subtask)
        print(psp)
        self.insert_subtask_into_database(project_name,subtask,psp)
    def insert_subtask_into_database(self,project,subtask,psp):
        created_by=self.label_3.text()
        try:
            conn=mysql.connector.connect(
                host="10.95.136.128",
                user="app_user", 
                password="Mohammed&meeraj786", 
                database="timesheet",
            )
            
            cursor=conn.cursor()
            query1="select id,project_name from projects where pspelement=%s and created_by=%s"
            cursor.execute(query1,(psp,created_by))
            projects=cursor.fetchall()
            if not projects:
                QMessageBox.warning(self,"No Project","Project not found!")
                return
            
            # project_id=v[0]
            # print("The project id is ",project_id)
            query2="insert into subtasks (project_id,project_name,pspelement,subtask,created_by) values(%s,%s,%s,%s,%s)"
            for project_id,project_name in projects:
                if project_name==project:
                    values=(project_id,project_name,psp,subtask,created_by)
                    cursor.execute(query2,values)
            # values=(project_id,project,psp,subtask,created_by)
            # cursor.execute(query2,values)
            conn.commit()
            print("Subtasks saved successfully")
            self.fetch_subtask_data()
            



            
        except mysql.connector.Error as err:
            print("Error occured",err)
        except Exception as ex:
            print("An exception occured")
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
    def fetch_subtask_data(self):
        created_by=self.label_3.text()
        try:
            conn=mysql.connector.connect(
                host="10.95.136.128",
                user="app_user", 
                password="Mohammed&meeraj786", 
                database="timesheet",
            )
            cursor=conn.cursor()
            query="select project_id,project_name,pspelement,subtask,created_date from subtasks where created_by=%s"
            cursor.execute(query,(created_by,))
            results=cursor.fetchall()
            self.populate_subtask_table(results)
        except mysql.connector.Error as err:
            print("Error occured",err)
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
    def insert_assigned_projects(self,assignment_ob,username,project,subtask,start_date,end_date,duration):
        print(username,project,subtask,start_date,end_date,duration)
        assigned_by=self.label_3.text()
        try:
            conn=mysql.connector.connect(
                host="10.95.136.128",
                user="app_user",
                password="Mohammed&meeraj786",
                database="timesheet"





            )
            cursor=conn.cursor(buffered=True)
            query="select user_id from users where username=%s"
            cursor.execute(query,(username,))
            u=cursor.fetchone()
            user_id=u[0]
            query="select id from projects where project_name=%s"
            cursor.execute(query,(project,))
            p=cursor.fetchone()
            project_id=p[0]
            query="select subtask_id from subtasks where subtask=%s"
            cursor.execute(query,(subtask,))
            s=cursor.fetchone()
            subtask_id=s[0]
            query = """
    INSERT INTO assignments 
    (user_id, project_id, subtask_id, username, project, subtask, start_date, end_date, duration,assigned_by) 
    VALUES (%s, %s, %s, %s, %s, %s, STR_TO_DATE(%s, '%d-%m-%Y'), STR_TO_DATE(%s, '%d-%m-%Y'), %s,%s)
"""

            values=(user_id,project_id,subtask_id,username,project,subtask,start_date,end_date,duration,assigned_by)
            cursor.execute(query,values)
            conn.commit()
            if cursor.rowcount>0:
                # assignment_ob.accept()
                pass
            self.fetch_assigned_projects()
        except mysql.connector.Error as err:
            print("Error occured while inserting",err)
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()

    
    def fetch_assigned_projects(self):
        assigned_by=self.label_3.text()
        try:
            conn=mysql.connector.connect(
                host="10.95.136.128",
                user="app_user",
                password="Mohammed&meeraj786",
                database="timesheet"
            )
            cursor=conn.cursor()
            query="Select assignment_id,user_id,username,project,subtask,start_date,end_date,duration from assignments where assigned_by=%s"
            cursor.execute(query,(assigned_by,))
            data=cursor.fetchall()
            self.populate_table(data,self.assignment_table)
        except  mysql.connector.Error as err:
            QMessageBox.critical(self,"Error occured while fetching",f"{err}")

    def populate_subtask_table(self,data):
        self.subtask_table.setRowCount(len(data))
        for row_idx,row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                item=QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                self.subtask_table.setItem(row_idx,col_idx,item)
        self.subtask_table.resizeColumnsToContents()

    def user_assignment_form(self):
        if self.open_user_assignment_form:
            self.open_user_assignment_form.populate_user_combo_box()
        else:
            self.open_user_assignment_form=AssignUser(self.get_project_names,self.label_3.text())
            print("i am being hit for some reasson")
        projects=self.get_project_names()
        self.open_user_assignment_form.update_project_names(project_names=projects)
        self.open_user_assignment_form.form_data_submitted.connect(self.insert_assigned_projects)
        self.open_user_assignment_form.exec()


    
