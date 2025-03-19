from ui.py.employee_dashboard_2 import Ui_MainWindow
from PyQt6.QtWidgets import QApplication,QMainWindow,QPushButton,QTableWidgetItem,QMessageBox,QComboBox,QFileDialog,QCalendarWidget
import sys,subprocess,os,socket,platform,uuid
from db.db import GetCursor
from datetime import datetime, timedelta
from openpyxl import Workbook,load_workbook
from PyQt6.QtCore import QEventLoop,QSizeF,QSettings,Qt
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from db.database_worker import DatabaseWorker
from .excel_user_report import Extract_to_Excel
import mysql.connector
import pandas as pd
import datetime as dt
from db.db_pool import DatabasePool
from .logged_in_users_employee import LoggedInUsersEmployee

from sqlalchemy.sql import text
from PyQt6 import QtGui
from fpdf import FPDF
class PDF(FPDF):
    def __init__(self,total_hours, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.column_widths = []
        self.page_width=210
        self.total_hours=total_hours
        self.page_margin=10
        self.max_table_width = self.page_width - 2 * self.page_margin
        self.table_start_y = 30 
    def resource_path(self,relative_path):
        if getattr(sys, '_MEIPASS', False):
                base_path = sys._MEIPASS
                # print(f"Running in bundled mode. Base path: {base_path}")
        else:
                current_dir=os.path.dirname(os.path.abspath(__file__))
                # base_path = os.path.abspath(os.path.join(current_dir, "..",".."))
                project_root = os.path.abspath(os.path.join(current_dir, ".."))
                base_path = os.path.join(project_root)
                
                # print(f"Running in source mode. Base pat: {base_path}")
        full_path = os.path.join(base_path, relative_path)
        # print(f"Resolved path for {relative_path}: {full_path}")
        return full_path
    def header(self):
        # Add company logo
        self.image(self.resource_path("assets/icons/shuecologo.png"), x=155, y=5, w=50)
        # Add user and week details
        self.set_font("Helvetica", size=10)
        self.set_xy(10, 15)
        
        self.ln(20)  # Add some spacing
        # Add total hours worked
        self.set_font("Helvetica", style="B", size=12)
        self.cell(0, 5, txt=f"Total Hours Worked: {self.total_hours}", ln=True, align="L")
        # self.ln(10)  # Add more spacing below header
        # Add a horizontal line
        self.set_draw_color(128, 128, 128)
        self.set_line_width(0.3)
        self.line(10, 40, 200, 40)

    def calculate_column_widths(self, headers, table,extra_width=5):
        """Calculate dynamic column widths based on content."""
        self.column_widths = []
        total_width=0
        for col in range(len(headers)):
            max_width = self.get_string_width(headers[col]) + 6  # Include padding
            for row in range(table.rowCount()):
                cell_widget = table.cellWidget(row, col)
                if isinstance(cell_widget, QComboBox):
                    content = cell_widget.currentText()
                else:
                    item = table.item(row, col)
                    content = item.text() if item else ""
                max_width = max(max_width, self.get_string_width(content) + 4)
            max_width+=extra_width
            self.column_widths.append(max_width)
            total_width+=max_width
        if total_width>self.max_table_width:
            scale_factor = self.max_table_width / total_width
            self.column_widths = [w * scale_factor for w in self.column_widths]
    
    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", size=8)
        self.cell(0, 10, "For Internal Circulation Only", align="C")
        


    def check_table_fit(self, y_start):
        """Check if there's enough space for the table or add a new page."""
        if self.get_y() + y_start > self.page_break_trigger:
            self.add_page()
    def draw_table(self, table,append_labels=None):
        headers = [table.horizontalHeaderItem(col).text() for col in range(table.columnCount())]
        self.calculate_column_widths(headers, table)
        self.set_fill_color(200, 200, 200)
        self.set_text_color(0, 0, 0)

        # Draw headers
        self.check_table_fit(self.table_start_y)  # Ensure headers fit on the page
        self.set_font("Helvetica", style="B", size=10)
        for i, header in enumerate(headers):
            self.cell(self.column_widths[i], 10, header, border=1, align="C")
        self.ln()

        # Draw table rows
        self.set_font("Helvetica", size=8)
        for row in range(table.rowCount()):
            self.check_table_fit(10)  # Ensure rows fit on the page
            fill = row % 2 == 0 
            for col in range(table.columnCount()):
                cell_widget = table.cellWidget(row, col)
                if isinstance(cell_widget, QComboBox):
                    content = cell_widget.currentText()
                else:
                    item = table.item(row, col)
                    content = item.text() if item else ""
                self.cell(self.column_widths[col], 10, content, border=1, align="C",fill=fill)
            self.ln()
        # if append_labels:
        #     self.check_table_fit(10)
        #     for col in range(len(headers)):
        #         if col >= 3:  # Start from the 4th column
        #             label_text = append_labels[col - 3] if col - 3 < len(append_labels) else ""
        #             self.cell(self.column_widths[col], 10, label_text, border=1, align="C")
        #         else:
        #             self.cell(self.column_widths[col], 10, "", border=1, align="C")
        #     self.ln()
        if append_labels:
            self.check_table_fit(10)

            # Merge the first three columns for "Total Time"
            merged_width = sum(self.column_widths[:3])  # Total width of the first three columns
            self.cell(merged_width, 10, "Total Time", border=1, align="C")

            # Add labels in the remaining columns
            for col in range(len(headers) - 3):
                label_text = append_labels[col] if col < len(append_labels) else ""
                self.cell(self.column_widths[col + 3], 10, label_text, border=1, align="C")
            self.ln()
        
class MyEmployee(QMainWindow,Ui_MainWindow):
    def __init__(self,username):
        super().__init__()
        self.username=username
        self.setupUi(self)
        self.setWindowTitle("Sidebar Menu")
        self.project_btn_1.setChecked(True)
        self.project_btn_2.setChecked(True)
        self.project_btn_1.clicked.connect(self.switch_to_projectPage)
        self.project_btn_2.clicked.connect(self.switch_to_projectPage)
        self.tableWidget.setRowCount(0)
        self.label_3.setText(username)
        self.setWindowTitle("Employee Dashboard")
        # self.update_icons()
        self.is_present=True
        self.threads=[]
        self.combo_box=QComboBox()
        self.combo_box.addItems(["-","IH-PD","BU Doors","BU Windows","BU Facades","PM","TD","DHI","BU Sliders","CE-V8 Process","CE - CIP","Fabrication","Testing","SIPL"])
        self.sign_out_btn_1.clicked.connect(self.show_login_window)
        self.sign_out_btn_2.clicked.connect(self.show_login_window)
        self.search_le.textChanged.connect(self.filter_table)
        self.logs_btn.clicked.connect(self.show_my_logs)
        self.regularization_btn_1.clicked.connect(self.show_regularization_page)
        self.regularization_btn_2.clicked.connect(self.show_regularization_page)
        self.project_btn_1.clicked.connect(self.show_project_page)
        self.project_btn_2.clicked.connect(self.show_project_page)
        self.stackedWidget.setCurrentIndex(0)
        self.btn_sign_in_sign_out.clicked.connect(lambda:(self.handle_logs()))
        

        # self.update_projects()
        self.week_combo.clear()
        self.week_combo.addItems([str(i) for i in range(1, 54)])
        self.week_combo.currentIndexChanged.connect(lambda:(self.update_table_headers_2(int(self.week_combo.currentText())),self.populate_table_by_week()))
        self.icon_name_widget.setHidden(True)
        self.update_table_headers_2(int(self.week_combo.currentText()))
        # self.populate_table_by_week()
        # self.cal_daily_weekly_totals()
        self.btn_add_project.clicked.connect(self.add_project_2)
        self.project_combo.currentIndexChanged.connect(self.update_subtasks)
        self.btn_save_calc.clicked.connect(lambda:(self.update_time_4()))
        self.btn_delete.clicked.connect(self.delete_timesheet_row)
        # self.export_btn.clicked.connect(lambda:self.export_to_excel(self.tableWidget))
        self.export_btn.clicked.connect(self.show_export_to_excel)
        self.btn_refresh.clicked.connect(self.refresh_data)
        self.reporting_to=""
        self.update_projects()
        self.update_subtasks()
        self.populate_table_on_state()
        self.get_manager()
        self.calendarWidget.setFirstDayOfWeek(Qt.DayOfWeek.Monday)
        self.calendarWidget.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)
        self.check_user_log()


        self.calendarWidget.setStyleSheet("""
QCalendarWidget QWidget {
    background-color: #fff;
    color: #000;
    border-radius: 2px;
    font-size: 14px;
}            

QCalendarWidget QAbstractItemView {
    background-color: #fff;
    color: #000;  /* Fixed text color */
    selection-background-color: #1F95EF;
}      

QCalendarWidget QTableView::item {
    border: 1px solid #f2f2f2; /* Border around individual cells */
}
QCalendarWidget QWidget#qt_calendar_navigationbar {
        background-color: #eee;
        
    } 
QCalendarWidget QToolButton {
                                          background-color: #eee;
    color: black;
    border-radius: 5px;
    padding: 15px;
                                          
                                          
                                          }     
                    
                                          
                                          """)
        
       


    def check_user_log(self):
        settings = QSettings("Schueco", "app")
        token=settings.value("sign_in","")
        print("The token issss ",token)
        db_instance=DatabasePool(pool_type="read")
        if token:
            print("yess")
            with db_instance.get_db_connection() as conn:
                cursor=conn.cursor()
                try:
                    query = "SELECT employee_name FROM user_logs WHERE token=%s AND token_expiry > NOW()"
                    cursor.execute(query,(token,))
                    result=cursor.fetchone()
                    print("The result isssssss ",result)
                    if result and result[0]==f"{self.label_3.text()}":
                        
                        self.btn_sign_in_sign_out.setText("Sign out")
                except mysql.connector.Error as err:
                    pass





    def show_regularization_page(self):
        self.stackedWidget.setCurrentIndex(2)
    def show_project_page(self):
        self.stackedWidget.setCurrentIndex(0)
    def show_my_logs(self):
        logs_dialog=LoggedInUsersEmployee(self.label_3.text())
        logs_dialog.exec()
    def refresh_data(self):
        self.update_projects()
        self.update_subtasks()

    def show_export_to_excel(self):
        export_diaolg=Extract_to_Excel()
        export_diaolg.form_data_submitted.connect(self.save_to_excel)
        export_diaolg.exec()

    def save_to_excel(self,obj,filter_by,start_month_week,end_month_week):
        user=self.label_3.text()
        query="Select week,username,project, pspelement,segment,subtask,logged_hours from tasks where"
        params={}
        month_to_week = {
        "january":  (1, 5), "february":  (5, 9), "march":  (9, 14), "april":  (14, 18),
        "may":  (18, 22), "june":  (22, 27), "july":  (27, 31), "august":  (31, 35),
        "september":  (36, 40), "october":  (40, 44), "november":  (44, 48), "december":  (49, 53)
    }
        if filter_by.lower()=="week":
            query += " week between :start_week and :end_week and username=:user"
            params["start_week"]=start_month_week
            params["end_week"]=end_month_week
            params["user"]=user

        elif filter_by.lower()=="month":
            start_month=start_month_week.lower()
            end_month=end_month_week.lower()
            if start_month in month_to_week and end_month in month_to_week:
                start_week=month_to_week[start_month][0]
                end_week=month_to_week[end_month][1]
                query +=" week between :start_week and :end_week and username=:user"
                params["start_week"]=start_week
                params["end_week"]=end_week
                params["user"]=user

            else:
                QMessageBox.warning(self,"Invalid Month","One or both selected months are invalid!")
                return
        elif filter_by.lower()=="--extract by month or week--":
            QMessageBox.warning(self,"No Selection","Please select month or week.")
            return
        worker=DatabaseWorker((text(query),params))
        worker.result_ready.connect(self.handle_employee_data)
        worker.error_occured.connect(self.handle_error)
        worker.finished.connect(lambda:self.threads.remove(worker))
        self.threads.append(worker)
        worker.start()
        obj.close()


    def handle_employee_data(self,data):
        if not data:
            QMessageBox.warning(self,"No Data","There is no data to export.")
            return
        headers=["Week","Employee Name","Project","PSP","Segment","Subtask","Logged Hours"]
        df=pd.DataFrame(data,columns=headers)
        file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_path:
            try:
                df.to_excel(file_path, index=False, engine="openpyxl")  # Save as Excel
                 # Load workbook and select active sheet
                wb = load_workbook(file_path)
                ws = wb.active

            # Set column widths
                column_widths = [10, 20, 25, 15, 20, 25, 15]  # Adjust as needed
                for col_num, width in enumerate(column_widths, start=1):
                    ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

            # Apply background color to headers
                header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                for cell in ws[1]:  # First row (headers)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")  # Center headers

            # Center-align all data
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            # Save workbook
                wb.save(file_path)

                QMessageBox.information(self, "Success", f"Data successfully exported to:\n{file_path}")
        
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save file:\n{str(e)}")



    def filter_table(self):
        search_text=self.search_le.text().lower()
        for row in range(self.tableWidget.rowCount()):
            show_row=False
            for col in range(self.tableWidget.columnCount()):
                item=self.tableWidget.item(row,col)
                if item and search_text in item.text().lower():
                    show_row=True
                    break
            self.tableWidget.setRowHidden(row,not show_row)
        

    def switch_to_projectPage(self):
        self.stackedWidget.setCurrentIndex(0)
    

    def show_login_window(self):
        # confirmation=QMessageBox.question(self,"Confirm sign out",f"Are you sure you want to sign out?",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
        # if confirmation != QMessageBox.StandardButton.Yes:
        #     return
        # self.show_login_dialog()
        from .login import Login
        confirmation=QMessageBox.question(self,"Confirm sign out",f"Are you sure you want to sign out?",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
        if confirmation !=  QMessageBox.StandardButton.Yes:
            return
        # Capture logout time
        # logout_time=dt.datetime.now().strftime("%H:%M:%S")
        # print("The log out time is ",logout_time)

        # #Remove login token from the database
        # db_instance=DatabasePool(pool_type="write")
        # with db_instance.get_db_connection() as conn:
        #     cursor=conn.cursor()
        #     try:
        #         query="update users set login_token=NULL, token_expiry=NULL where username=%s"
        #         cursor.execute(query,(self.label_3.text(),))
        #         print(f"The user is -{self.label_3.text()}-")

        #         query="select login_time from user_logs where employee_name=%s order by login_time desc limit 1"
        #         cursor.execute(query,(self.label_3.text(),))
        #         login_hour=cursor.fetchone()[0]
        #         print("the login hour is ",login_hour)
        #         total_logged_hours=dt.datetime.strptime(logout_time,"%H:%M:%S")-login_hour
        #         query="update user_logs set logout_time=%s,hours_logged=%s where employee_name=%s and login_time=%s"
        #         cursor.execute(query,(logout_time,total_logged_hours,self.label_3.text(),login_hour))

        #         conn.commit()
        #     except mysql.connector.Error as err:
        #         conn.rollback()
        #         QMessageBox.critical(self,"Database Error",f"An error occurred: {err}")
        
        settings=QSettings("Schueco","app")
        settings.remove("login_token")

        login_dialog=Login()
        self.close()
        login_dialog.exec()
        
    def show_login_dialog(self):
        from .login import Login
        login_dialog=Login()
        login_dialog.exec()
        self.close()
    def resource_path(self,relative_path):
        if getattr(sys, '_MEIPASS', False):
                base_path = sys._MEIPASS
                # print(f"Running in bundled mode. Base path: {base_path}")
        else:
                current_dir=os.path.dirname(os.path.abspath(__file__))
                # base_path = os.path.abspath(os.path.join(current_dir, "..",".."))
                project_root = os.path.abspath(os.path.join(current_dir, ".."))
                base_path = os.path.join(project_root)
                
                # print(f"Running in source mode. Base pat: {base_path}")
        full_path = os.path.join(base_path, relative_path)
        # print(f"Resolved path for {relative_path}: {full_path}")
        return full_path

    
    # def update_icons(self):
    #     icon1=QtGui.QIcon()
    #     icon1.addPixmap(QtGui.QPixmap(self.resource_path("assets/icons/log_out_white.png")), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
    #     icon1.addPixmap(QtGui.QPixmap(self.resource_path("assets/icons/log_out.png")), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.On)
    #     self.sign_out_btn_1.setIcon(icon1)
    #     self.sign_out_btn_2.setIcon(icon1)


    # def resource_path(self,relative_path):
    #     if getattr(sys, '_MEIPASS', False):
    #             base_path = sys._MEIPASS
    #             print(f"Running in bundled mode. Base path: {base_path}")
    #     else:
    #             base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
    #             print(f"Running in source mode. Base path: {base_path}")
    #     full_path = os.path.join(base_path, relative_path)
    #     print(f"Resolved path for {relative_path}: {full_path}")
    #     return full_path
    def export_to_excel(self,table):
        workbook=Workbook()
        sheet=workbook.active
        sheet.title="Exported Data"
        sheet.column_dimensions['A'].width=25
        sheet.column_dimensions['B'].width=20
        sheet.column_dimensions['C'].width=20
        sheet.column_dimensions['D'].width=10
        sheet.column_dimensions['E'].width=10
        sheet.column_dimensions['F'].width=10
        sheet.column_dimensions['G'].width=10
        sheet.column_dimensions['H'].width=10
        sheet.column_dimensions['I'].width=10
        sheet.column_dimensions['J'].width=10
        sheet.column_dimensions['k'].width=10
        last_row=1
        bold_font=Font(bold=True)
        fill=PatternFill(start_color="73a4ca",end_color="73a4ca",fill_type="solid")
        sheet.cell(row=1,column=1).value="Week"
        sheet.cell(row=1,column=1).font=bold_font
        
        current_week=self.week_combo.currentText()
        sheet.cell(row=1,column=2).value=current_week
        sheet.cell(row=1,column=2).font=bold_font
        headers=[]
        daily_totals = []  

        for i in range(7): 
            label_widget = getattr(self, f"day{i+1}_value")
            daily_totals.append(label_widget.text() if label_widget else 0)
        print("The daily totals are ",daily_totals)
        # this loop copies the headers to the list and sets the headers in excel
        for column in range(table.columnCount()):
            headers.append(table.horizontalHeaderItem(column).text())
            sheet.cell(row=2,column=column+1).value=table.horizontalHeaderItem(column).text()
            sheet.cell(row=2,column=column+1).font=bold_font
            sheet.cell(row=2,column=column+1).fill=fill
            sheet.cell(row=2,column=column+1).alignment=Alignment(horizontal="center")

        for row in range(table.rowCount()):
            last_row+=1
            for column in range(table.columnCount()):
                cell_widget=table.cellWidget(row,column)
                if isinstance(cell_widget,QComboBox):
                    
                    sheet.cell(row=row + 3, column=column + 1).value = cell_widget.currentText()
                    sheet.cell(row=row + 3, column=column + 1).alignment=Alignment(horizontal="center")
                else:
                    item=table.item(row,column)
                    sheet.cell(row=row + 3, column=column + 1).value=item.text() if item else ""
        
        print("last row is",last_row)
        for i,val in enumerate(daily_totals):
            if i == 0:
                sheet.merge_cells(start_row=last_row+2,start_column=i+1,end_row=last_row+2,end_column=i+3)
                sheet.cell(row=last_row+2,column=i+1).value="Total Time"
                sheet.cell(row=last_row+2,column=i+1).alignment=Alignment(horizontal="center")
                sheet.cell(row=last_row+2,column=i+1).font=bold_font
                sheet.cell(row=last_row+2,column=i+1).fill=fill
                sheet.cell(row=last_row+2,column=i+1).alignment=Alignment(horizontal="center")
            sheet.cell(row=last_row+2,column=i+4).value=val
            if i==len(daily_totals)-1:
                sheet.cell(row=last_row+2,column=i+5).value=self.total_time.text()
                sheet.cell(row=last_row+2,column=i+5).font=bold_font
                sheet.cell(row=last_row+2,column=i+5).alignment=Alignment(horizontal="center")
        
        msg_box=QMessageBox()
        msg_box.setWindowTitle("Export Options")
        msg_box.setText("Choose the export format:")
        excel_button=msg_box.addButton("Export as Excel",QMessageBox.ButtonRole.AcceptRole)
        pdf_button = msg_box.addButton("Export as PDF", QMessageBox.ButtonRole.AcceptRole)
        cancel_button = msg_box.addButton(QMessageBox.StandardButton.Cancel)
        msg_box.exec()
        if msg_box.clickedButton()==excel_button:
            options=QFileDialog.Option.DontUseNativeDialog
            file_path,_=QFileDialog.getSaveFileName(None,"Save Excel File","","Excel Files(*.xlsx)",options=options)
            if file_path:
                if not file_path.endswith(".xlsx"):
                    file_path+=".xlsx"
                workbook.security.workbookPassword="pass123"
                workbook.security.lockStructure=True
                for sheet in workbook.worksheets:
                    sheet.protection.set_password("pass123")
                    sheet.protection.enable()
                workbook.save(file_path)
                print("success")
        elif msg_box.clickedButton()==pdf_button:
            options = QFileDialog.Option.DontUseNativeDialog
            file_path, _ = QFileDialog.getSaveFileName(None, "Save PDF File", "", "PDF Files(*.pdf)", options=options)
            if file_path:
                if not file_path.endswith(".pdf"):
                    file_path += ".pdf"
                self.export_to_pdf(table, file_path)
                # self.export_to_pdf_2()
                print("PDF file exported successfully.")
        else:
            pass

   

    def export_to_pdf(self,table,file_path):
        total_hours=self.total_time.text()

        pdf = PDF(total_hours=total_hours)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_font("Arial", style="B", size=14)
        # pdf.cell(0, 10, txt="Exported Data", ln=True, align="C")
        pdf.image(self.resource_path("assets/icons/shuecologo.png"), x=155, y=5, w=50)
        pdf.ln(60)
        user=self.label_3.text()
        week=self.week_combo.currentText()
        pdf.cell(0, 10, txt=f"User: {user}", ln=True, align="L")
        pdf.cell(0, 10, txt=f"Week: {week}", ln=True, align="L")
        # pdf.cell(0, 10, txt=f"Total Hours Worked: {total_hours}", ln=True, align="L")
        pdf.ln(5)
        daily_totals=[]
        for i in range(7): 
            label_widget = getattr(self, f"day{i+1}_value")
            daily_totals.append(label_widget.text() if label_widget else 0)
        
        pdf.draw_table(table,daily_totals)

        pdf.output(file_path)
        print("pdf saved successfully")

    def export_to_pdf_2(self):
        pdf_writer=QtGui.QPdfWriter("table_output.pdf")
        pdf_writer.setPageSize(QtGui.QPageSize(QSizeF(210, 297), QtGui.QPageSize.SizeMode.MillimeterMode))
        painter=QtGui.QPainter(pdf_writer)
        painter.begin(pdf_writer)
        painter.setFont(QtGui.QPainter.font())
        rows=self.tableWidget.rowCount()
        cols=self.tableWidget.columnCount()
        row_height=20
        column_width=100
        x, y = 50, 50
        for row in range(rows):
            for col in range(cols):
                item = self.table.item(row, col)
                if item:
                    painter.drawText(x + col * column_width, y + row * row_height, item.text())
                else:
                    widget = self.table.cellWidget(row, col)
                    if isinstance(widget, QComboBox):
                        painter.drawText(x + col * column_width, y + row * row_height, widget.currentText())

        label_x = x + cols * column_width
        label_y = y + rows * row_height
        painter.drawText(label_x, label_y, "Total:")
        painter.end()
        
    def delete_timesheet_row(self):
        selected_row=self.tableWidget.currentRow()
        if selected_row==-1:
            QMessageBox.warning(self,"No Selection","Please select a row to delete.")
            return
        project=self.tableWidget.item(selected_row,0).text()
        subtask=self.tableWidget.item(selected_row,2).text()
        segment=self.tableWidget.cellWidget(selected_row,1).currentText()
        user=self.label_3.text()
        week=self.week_combo.currentText()
        confirmation=QMessageBox.question(self,"Confirm deletion",f"Are you sure you want to delete?",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
        if confirmation != QMessageBox.StandardButton.Yes:
            return
        try:
            conn=mysql.connector.connect(
            # host="127.0.0.1",
            # user="Meeraj", 
            # password="Mohammed&meeraj786", 
            # database="timesheet",
            # port=3306
            host="10.95.136.128",
            user="app_user", 
            password="Mohammed&meeraj786", 
            database="timesheet",


            )
            cursor=conn.cursor()
            query="delete from tasks where project=%s and subtask=%s and username=%s and week=%s and segment=%s"
            values=(project,subtask,user,week,segment)
            cursor.execute(query,values)
            conn.commit()
            self.tableWidget.removeRow(selected_row)
            self.cal_daily_weekly_totals()
            self.update_time_4()
            self.insert_weekly_time()
            
        except mysql.connector.Error as err:
            QMessageBox.critical(self,"Failed",f"{err.msg}")
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
    # def closeEvent(self, event):
        
    #     reply=QMessageBox.question(self,"Confirm Exit","Are you sure you want to close the app?",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
    #     if reply==QMessageBox.StandardButton.Yes:
    #         self.on_close()
    #         event.accept()
    #     else:
    #         event.ignore()


    def on_close(self):
        user=self.label_3.text()
        week=int(self.week_combo.currentText())
        query=text("insert into user_state (username,week) values(:user,:week) on duplicate key update week=:week")
        params={"user":user,"week":week}
        worker1=DatabaseWorker((query,params))
        worker1.set_write_operation(True)
        worker1.result_ready.connect(lambda _:None)
        worker1.finished.connect(lambda:self.threads.remove(worker1))
        worker1.error_occured.connect(self.handle_error)
        self.threads.append(worker1)
        worker1.start()

    def get_login_details(self):
        # login_time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        login_time=dt.datetime.now().strftime("%H:%M:%S")
        login_date=dt.datetime.now().strftime("%Y-%m-%d")
        date_obj=dt.datetime.strptime(login_date,"%Y-%m-%d")
        day_name=date_obj.strftime("%A")
        try:
            hostname=socket.gethostname()
            ipaddress=socket.gethostbyname(hostname)
        except Exception as e:
            ipaddress=f"Error: {e}"
            print(ipaddress)
        #Get Wifi Name
        wifi_name="Unknown"
        if platform.system()=="Windows":
            try:
                result=subprocess.run(["netsh", "wlan", "show", "interfaces"], capture_output=True, text=True)
                for line in result.stdout.split("\n"):
                    if "SSID" in line:
                        wifi_name=line.split(":")[1].strip()
                        break
            except Exception as e:
                wifi_name=f"Error: {e}"
        return login_time,ipaddress,wifi_name,login_date,day_name


    def handle_logs(self):
        if self.btn_sign_in_sign_out.text()=="Sign In":
            token=str(uuid.uuid4())
            expiration_time=dt.datetime.now().replace(microsecond=0)+dt.timedelta(hours=10)
            
            username=self.label_3.text()
            login_time,ipaddress,wifi_name,login_date,day_name=self.get_login_details()
            system_user=os.getlogin()
            entry_exists=self.push_logs_to_db(login_time,wifi_name,login_date,day_name,system_user,username,token,expiration_time)
            if entry_exists:
                return
            QMessageBox.information(self,"Success","Logged in succesfully.")
            self.btn_sign_in_sign_out.setText("Sign Out")
            # self.btn_sign_in_sign_out.setStyleSheet("background-color: red")
            # self.btn_sign_in_sign_out.setIcon(QtGui.QIcon(self.resource_path("assets/icons/sign_out.png")))
            # self.btn_sign_in_sign_out.setIconSize(QtCore.QSize(20,20))
        else:
            message=QMessageBox.question(self,"Confirm sign out","Are you sure you want to sign out?",QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
            if message==QMessageBox.StandardButton.Yes:
                self.capture_log_out_time()
                self.btn_sign_in_sign_out.setText("Sign In")
                settings=QSettings("Schueco","app")
                settings.remove("sign_in")

            


    def capture_log_out_time(self):
        logout_time=dt.datetime.now().strftime("%H:%M:%S")
        print("The log out time is ",logout_time)

        #Remove login token from the database
        db_instance=DatabasePool(pool_type="write")
        with db_instance.get_db_connection() as conn:
            cursor=conn.cursor()
            try:
                query="update users set login_token=NULL, token_expiry=NULL where username=%s"
                cursor.execute(query,(self.label_3.text(),))
                print(f"The user is -{self.label_3.text()}-")

                query="select login_time from user_logs where employee_name=%s order by login_time desc limit 1"
                cursor.execute(query,(self.label_3.text(),))
                login_hour=cursor.fetchone()[0]
                print("the login hour is ",login_hour)
                total_logged_hours=dt.datetime.strptime(logout_time,"%H:%M:%S")-login_hour
                query="update user_logs set logout_time=%s,hours_logged=%s where employee_name=%s and login_time=%s"
                cursor.execute(query,(logout_time,total_logged_hours,self.label_3.text(),login_hour))

                conn.commit()
            except mysql.connector.Error as err:
                conn.rollback()
                QMessageBox.critical(self,"Database Error",f"An error occurred: {err}")
        
        settings=QSettings("Schueco","app")
        settings.remove("login_token")
    def push_logs_to_db(self,login_time,wifi_name,login_date,day_name,system_user,username,token,expiration_time):
        db_instance=DatabasePool(pool_type="write")
        with db_instance.get_db_connection() as conn:
            cursor=conn.cursor()
            try:
                query="select user_id from users where binary username=%s"
                cursor.execute(query,(username,))
                user_id=cursor.fetchone()[0]
                query="insert into user_logs(employee_id,employee_name,username,login_time,wifi_name,login_date,day_name,token,token_expiry) values(%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                cursor.execute(query,(user_id,username,system_user,login_time,wifi_name,login_date,day_name,token,expiration_time))
                settings=QSettings("Schueco","app")
                settings.setValue("sign_in",token)
                conn.commit()
            except mysql.connector.errors.IntegrityError as e:
                if e.errno==1062:
                    QMessageBox.warning(self,"Duplicate Entry","You have already signed in.")
                    return True
                else:
                    print(f"Database error")
            except mysql.connector.Error as err:
                conn.rollback()
                # pass
                QMessageBox.critical(self,"Database Error",f"An error occured: {err}")


    # def update_time(self):
    #     # is_valid=self.cal_daily_weekly_totals()
    #     # if not is_valid:
    #     #     QMessageBox.warning(self,"Invalid","The total hours must not exceed 8")
    #     #     return
    #     for row in range(self.tableWidget.rowCount()):
    #         project=self.tableWidget.item(row,0).text()
    #         subtask=self.tableWidget.item(row,1).text()
    #         week=int(self.week_combo.currentText())
    #         user=self.label_3.text()
    #         daily_hours=[]
    #         for col in range(2,self.tableWidget.columnCount()):
    #             item=self.tableWidget.item(row,col)
    #             hours=int(item.text()) if item and item.text().isdigit() else 0
    #             if hours>8:
    #                 QMessageBox.warning(self,"Input Error","Hours cannot be more than 8.")
    #                 return
                    
    #             daily_hours.append(hours)
    #         db=GetCursor()
    #         if db.conn and db.cursor:
    #             query="update tasks set weekday_1=%s,weekday_2=%s,weekday_3=%s,weekday_4=%s,weekday_5=%s,weekday_6=%s,weekday_7=%s where project=%s and subtask=%s and week=%s and username=%s"
    #             db.cursor.execute(query,(*daily_hours,project,subtask,week,user))
    #             db.conn.commit()

    #         db.close_connection()
    #     print("updated successfully")
        
    #     # Calculate daily totals
    #     self.cal_daily_weekly_totals()
    #     self.insert_weekly_time()

    
    
    # def update_time_3(self):
    #     user = self.label_3.text()
    #     week = int(self.week_combo.currentText())
    #     db = GetCursor()
    #     if db.conn and db.cursor:
    #         try:
    #             for row in range(self.tableWidget.rowCount()):
    #                 project_id = self.tableWidget.item(row, 0).text()
    #                 segment = self.tableWidget.cellWidget(row, 1).currentText()
    #                 subtask = self.tableWidget.item(row, 2).text()

    #                 query = "SELECT COUNT(*) FROM tasks WHERE username=%s AND week=%s AND project=%s AND segment=%s AND subtask=%s"
    #                 values = (user, week, project_id, segment, subtask)
    #                 db.cursor.execute(query, values)
    #                 is_exists = db.cursor.fetchone()[0] 
    #                 daily_hours = []
    #                 for col in range(3, self.tableWidget.columnCount()):
    #                     item = self.tableWidget.item(row, col)
    #                     hours = int(item.text()) if item and item.text().isdigit() else 0
    #                     if hours > 8:
    #                         QMessageBox.warning(self, "Input Error", "Hours cannot be more than 8.")
    #                         return
    #                     daily_hours.append(hours)
    #                 print(hours)
    #                 if not is_exists:
    #                     query = """
    #                     INSERT INTO tasks
    #                     (week, username, project, segment, subtask, weekday_1, weekday_2, weekday_3, weekday_4, weekday_5, weekday_6, weekday_7)
    #                     VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    #                 """
    #                     values = (week, user, project_id, segment, subtask, *daily_hours)
    #                 else:
    #                     query = """
    #                     UPDATE tasks
    #                     SET segment=%s, weekday_1=%s, weekday_2=%s, weekday_3=%s, weekday_4=%s, weekday_5=%s, weekday_6=%s, weekday_7=%s
    #                     WHERE project=%s AND subtask=%s AND week=%s AND username=%s
    #                 """
    #                     values = (segment, *daily_hours, project_id, subtask, week, user)
    #                     db.cursor.execute(query, values)
    #                 db.cursor.execute(query, values)
    #             db.conn.commit() 
    #             self.cal_daily_weekly_totals()
    #         except Exception as e:
    #             QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")
    #         finally:
    #             db.close_connection()
    #             print("success")


    def update_time_4(self):
        user = self.label_3.text()
        week = int(self.week_combo.currentText())
        row_data_set=set()
        reporting_to=self.reporting_to
        for row in range(self.tableWidget.rowCount()):
            project_id = self.tableWidget.item(row, 0).text()
            segment_combo = self.tableWidget.cellWidget(row, 1)
            segment=segment_combo.currentText() 
            # segment_combo.setEnabled(False)
            subtask = self.tableWidget.item(row, 2).text()
            row_data = (project_id, subtask, segment)
            if row_data in row_data_set:
                QMessageBox.warning(self,"Duplicate Rows",f"Duplicate found in row {row+1}:{row_data}")
                return
            else:
                row_data_set.add(row_data)


        db = GetCursor()
        if db.conn and db.cursor:
            try:
                for row in range(self.tableWidget.rowCount()):
                    project_id = self.tableWidget.item(row, 0).text()
                    pspelement=self.get_psp_element(project_id)
                    segment_combo = self.tableWidget.cellWidget(row, 1)
                    segment=segment_combo.currentText()
                    segment_combo.setEnabled(False)
                    subtask = self.tableWidget.item(row, 2).text()

                # Validate daily hours
                    daily_hours = []
                    for col in range(3, self.tableWidget.columnCount()):
                        item = self.tableWidget.item(row, col)
                        # hours = float(item.text()) if item and item.text().isdigit() else 0
                        # if hours > 8:
                        #     QMessageBox.warning(self, "Input Error", "Hours cannot be more than 8.")
                        #     return
                        if item and item.text().strip():
                            try:
                                hours=float(item.text().strip())
                                if hours>8:
                                    QMessageBox.warning(self,"Input Error","Hours cannot be more than 8.")
                                    return
                            except ValueError:
                                QMessageBox.warning(self,"Input Error",f"Invalid input for hours: {item.text()}")
                                return
                        else:
                            hours=0
                        daily_hours.append(hours)
                    print("The daily hours are ",daily_hours)
                    logged_hours=sum(daily_hours)
                # Upsert query
                    query = """
                    INSERT INTO tasks
                    (week, username, project, segment, subtask, weekday_1, weekday_2, weekday_3, weekday_4, weekday_5, weekday_6, weekday_7,reporting_to,pspelement,logged_hours)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE
                    segment = VALUES(segment),
                    weekday_1 = VALUES(weekday_1),
                    weekday_2 = VALUES(weekday_2),
                    weekday_3 = VALUES(weekday_3),
                    weekday_4 = VALUES(weekday_4),
                    weekday_5 = VALUES(weekday_5),
                    weekday_6 = VALUES(weekday_6),
                    weekday_7 = VALUES(weekday_7),
                    reporting_to=VALUES(reporting_to),
                    pspelement=VALUES(pspelement),
                    logged_hours=VALUES(logged_hours)
                """
                    values = (week, user, project_id, segment, subtask, *daily_hours,reporting_to,pspelement,logged_hours)

                    db.cursor.execute(query, values)
                    # logged_hours=sum(daily_hours)
                    # query="insert into tasks (logged_hours) values(%s) on duplicate key update logged_hours=values(logged_hours)"
                    # values=(logged_hours,)
                    # db.cursor.execute(query,values)

                db.conn.commit()  # Commit all changes
                self.cal_daily_weekly_totals()
                self.insert_weekly_time()

            except Exception as e:
                QMessageBox.critical(self, "Database Error", f"An error occurred: {e}")
            finally:
                db.close_connection()
                print("success")
    
    # def update_time_4(self):
    #     user=self.label_3.text()
    #     week=int(self.week_combo.currentText())

    #     queries=[]
    #     try:
    #         for row in range(self.tableWidget.rowCount()):
    #             project_id = self.tableWidget.item(row, 0).text()
    #             segment_combo = self.tableWidget.cellWidget(row, 1)
    #             segment=segment_combo.currentText()
    #             segment_combo.setEnabled(False)
    #             subtask = self.tableWidget.item(row, 2).text()

    #             #validate daily hours
    #             daily_hours=[]
    #             for col in range(3,self.tableWidget.columnCount()):
    #                 item=self.tableWidget.item(row,col)
    #                 hours=int(item.text()) if item and item.text().isdigit() else 0
    #                 if hours>8:
    #                     QMessageBox.warning(self,"Input Error", "Hours cannot be more than 8.")
    #                     return
    #                 daily_hours.append(hours)

    #             query=text("INSERT INTO tasks (week, username, project, segment, subtask, weekday_1, weekday_2, weekday_3, weekday_4, weekday_5, weekday_6, weekday_7) values(:week,:username,:project,:segment,:subtask,:weekday_1,:weekday_2,:weekday_3,:weekday_4,:weekday_5,:weekday_6,:weekday_7,) on duplicate key update segment=values(segment),weekday_1 = VALUES(weekday_1),weekday_2 = VALUES(weekday_2),weekday_3 = VALUES(weekday_3),weekday_4 = VALUES(weekday_4),weekday_5 = VALUES(weekday_5),weekday_6 = VALUES(weekday_6),weekday_7 = VALUES(weekday_7)")
    #             params={"week":week,
    #                     "username": user,
    #                     "project": project_id,
    #                     "segment": segment,
    #                     "subtask": subtask,
    #                     "weekday_1": daily_hours[0],
    #                     "weekday_2": daily_hours[1],
    #                     "weekday_3": daily_hours[2],
    #                     "weekday_4": daily_hours[3],
    #                     "weekday_5": daily_hours[4],
    #                     "weekday_6": daily_hours[5],
    #                     "weekday_7": daily_hours[6],
                        
                        
                        
                        
                        
                        
                        
    #                     }
    #             queries.append((query,params))
    #         def handle_success(_):
    #             self.cal_daily_weekly_totals()
    #             self.insert_weekly_time()
    #         worker=DatabaseWorker(queries)
    #         worker.set_write_operation(True)
    #         worker.result_ready.connect(handle_success)
    #         worker.error_occured.connect(self.handle_error)
    #         # self.threads.remove(lambda:)
    #         worker.finished.connect(lambda:self.threads.remove(worker))
    #         self.threads.append(worker)
    #         worker.start()



                 

    #     except Exception as e:
    #         QMessageBox.critical(self,"Error",f"An unexpected error occured: {e}")

    
        
    


                    


    




        
    
    def get_psp_element(self,project_id):
        query=text("select pspelement from projects where project_name=:project_name")
        params={"project_name":project_id}
        worker=DatabaseWorker((query,params))
        psp="-"
        loop=QEventLoop()
        def handle_data(data):
            nonlocal psp
            if data:
                psp=data[0][0]
            loop.quit()
        def handle_error(error):
            loop.quit()
            QMessageBox.warning(self,"Error",f"{error}")
        def cleanup():
            loop.quit()
            


        worker.result_ready.connect(handle_data)
        worker.error_occured.connect(handle_error)
        worker.finished.connect(cleanup)
        worker.finished.connect(lambda:self.threads.remove(worker))
        self.threads.append(worker)
        worker.start()
        loop.exec()
        return psp

    def cal_daily_weekly_totals(self):
        flag=True
        daily_totals=[0]*(self.tableWidget.columnCount()-3)
        for col in range(3,self.tableWidget.columnCount()):
            for row in range(self.tableWidget.rowCount()):
                item=self.tableWidget.item(row,col)
                # if item and item.text().isdigit():
                #     daily_totals[col-3]+=float(item.text())
                if item and item.text().strip():
                    try:
                        hours=float(item.text())
                        daily_totals[col-3]+=hours
                    except ValueError:
                        QMessageBox.warning(self,"Input Error",f"Invalid input for hours")
                        
        for i,total in enumerate(daily_totals):
            label=getattr(self,f"day{i+1}_value")
            if total>8:
                label.setText("Invalid")
                flag=False
            else:
                label.setText(str(total))
        #calculate weekly total and update label
        weekly_total=sum(daily_totals)
        self.weekly_total=weekly_total
        t=str(weekly_total)
        if not flag:
            self.total_time.setText("Invalid")
            self.weekly_total="Invalid"

        else:
            self.total_time.setText(f"{t} hours")

        return flag
    
        # insert total weekly hours
    def show_warning_message(self,warning):
        msgb=QMessageBox()
        msgb.setIcon(QMessageBox.Icon.Warning)
        msgb.setWindowTitle("Invalid input")
        msgb.setText(warning)
        msgb.setStandardButtons(QMessageBox.StandardButton.Ok)
        msgb.exec()
    # def insert_weekly_time(self):
    #     db=GetCursor()
    #     if db.conn and db.cursor:
    #         user=self.label_3.text()
    #         query="select reporting_to from users where username=%s"
    #         db.cursor.execute(query,(user,))
    #         reporting_to=db.cursor.fetchone()[0]
    #         week=int(self.week_combo.currentText())
    #         get_user_query="select user_id from users where username=%s"
    #         db.cursor.execute(get_user_query,(user,))
    #         user_result=db.cursor.fetchone()

    #         if not user_result:
    #             QMessageBox.critical(None, "Error","User not found in the database")
    #             return
    #         user_id=user_result[0]
    #         check_query="select count(*) from weekly_hours where username=%s and week=%s"
    #         db.cursor.execute(check_query,(user,week))
    #         row_exists=db.cursor.fetchone()[0]
    #         if row_exists>0:
    #             update_query="update weekly_hours set total_hours = %s where week=%s and username=%s"
    #             db.cursor.execute(update_query,(self.weekly_total,week,user))
    #         else:
    #             insert_query="insert into weekly_hours(user_id,week,username,total_hours,reporting_to) values(%s,%s,%s,%s,%s)"
    #             db.cursor.execute(insert_query,(user_id,week,user,self.weekly_total,reporting_to))
    #             print("inserted successfully")
    #         db.conn.commit()
    #     db.close_connection()
    def get_manager(self):
        user=self.label_3.text()
        query=text("select reporting_to from users where username=:user")
        params={"user":user}
        worker1=DatabaseWorker((query,params))
        def handle_reporting_to(result):
            # if not result:
            #     QMessageBox.critical(None,"Error","User not found in the database.")
            #     return
            self.reporting_to=result[0][0]
        worker1.result_ready.connect(handle_reporting_to)
        worker1.error_occured.connect(self.handle_error)
        worker1.finished.connect(lambda:self.threads.remove(worker1))
        self.threads.append(worker1)
        worker1.start()


    def insert_weekly_time(self):
        user=self.label_3.text()
        week=int(self.week_combo.currentText())
        query=text("select reporting_to from users where username=:user")
        params={"user":user}
        worker1=DatabaseWorker((query,params))

        def handle_reporting_to(result):
            if not result:
                QMessageBox.critical(None,"Error","User not found in the database.")
                return
            reporting_to=result[0][0]
            # Fetch user id
            query2=text("select user_id from users where username=:user")
            params2={"user":user}
            worker2=DatabaseWorker((query2,params2))

            def handle_user_id(result):
                if not result:
                    QMessageBox.critical(None,"Error","User id not found in the database")
                    return
                user_id=result[0][0]
                query3=text("select count(*) from weekly_hours where username=:user and week=:week")
                params3={"user":user,"week":week}
                worker3=DatabaseWorker((query3,params3))
                def handle_check_existence(result):
                    row_exists=result[0][0]
                    if row_exists>0:
                        update_query=text("update weekly_hours set total_hours=:total_hours where week=:week and username=:user")
                        update_params={"total_hours":self.weekly_total,"week":week,"user":user}
                        update_worker=DatabaseWorker((update_query,update_params))
                        update_worker.set_write_operation(True)
                        update_worker.result_ready.connect(lambda _:None)
                        update_worker.error_occured.connect(self.handle_error)
                        update_worker.finished.connect(lambda:self.threads.remove(update_worker))
                        self.threads.append(update_worker)
                        update_worker.start()
                    else:
                        insert_query=text("insert into weekly_hours (user_id, week, username, total_hours, reporting_to) values (:user_id, :week, :user, :total_hours, :reporting_to) ")
                        insert_params={"user_id":user_id,"week":week,"user": user,"total_hours": self.weekly_total,"reporting_to": reporting_to}
                        insert_worker=DatabaseWorker((insert_query,insert_params))
                        insert_worker.set_write_operation(True)
                        insert_worker.result_ready.connect(lambda _:None)
                        insert_worker.error_occured.connect(self.handle_error)
                        insert_worker.finished.connect(lambda:self.threads.remove(insert_worker))
                        self.threads.append(insert_worker)
                        insert_worker.start()

                worker3.result_ready.connect(handle_check_existence)
                worker3.error_occured.connect(self.handle_error)
                worker3.finished.connect(lambda:self.threads.remove(worker3))
                self.threads.append(worker3)
                worker3.start()

            worker2.result_ready.connect(handle_user_id)
            worker2.error_occured.connect(self.handle_error)
            worker2.finished.connect(lambda:self.threads.remove(worker2))
            self.threads.append(worker2)
            worker2.start()

        worker1.result_ready.connect(handle_reporting_to)
        worker1.error_occured.connect(self.handle_error)
        worker1.finished.connect(lambda:self.threads.remove(worker1))
        self.threads.append(worker1)
        worker1.start()
        
        

    # def add_project(self):
    #     week=self.week_combo.currentText()
    #     project=self.project_combo.currentText()
    #     subtask=self.subtask_combo.currentText()
    #     user=self.label_3.text()
    #     # print(user,project,subtask,week)
    #     db=GetCursor()
    #     if db.conn and db.cursor:
    #         query="select * from tasks where week=%s and project=%s and subtask=%s and username=%s"
    #         values=(week,project,subtask,user)
    #         db.cursor.execute(query,values)
    #         data_exists=db.cursor.fetchone()
    #         # print(data_exists)
    #         if data_exists:
    #             pass
    #         else:
    #             query="insert into tasks (week,username,project,subtask) values(%s,%s,%s,%s)"
    #             values=(week,user,project,subtask)
    #             try:
    #                 db.cursor.execute(query,values)
    #                 db.conn.commit()
    #                 self.add_to_table(project,subtask)
    #             except Exception as e:
    #                 QMessageBox.warning(self,"Failed",f"Database error {e}")
    #     db.close_connection()

    def add_project_2(self):
        project=self.project_combo.currentText()
        subtask=self.subtask_combo.currentText()
        self.add_to_table_2(project,subtask)

          
    def add_to_table_2(self,project,subtask):
        row_count=self.tableWidget.rowCount()
        self.tableWidget.insertRow(row_count)
        combo_box=QComboBox()
        combo_box.addItems(["-","IH-PD","BU Doors","BU Windows","BU Facades","PM","TD","DHI","BU Sliders","CE-V8 Process","CE - CIP","Fabrication","Testing","SIPL"])
        
        self.tableWidget.setItem(row_count,0,QTableWidgetItem(project))
        self.tableWidget.setCellWidget(row_count,1, combo_box)

        self.tableWidget.setItem(row_count,2,QTableWidgetItem(subtask))
        combo_box.currentIndexChanged.connect(lambda:self.handle(row_count,combo_box))
    def handle(self,row,combo_box):
        print("hello boys")
        project_id = self.tableWidget.item(row, 0).text()
        subtask = self.tableWidget.item(row, 2).text()
        selected_segment = combo_box.currentText()
        for r in range(self.tableWidget.rowCount()):
            if r != row:
                existing_project = self.tableWidget.item(r, 0).text()
                existing_subtask = self.tableWidget.item(r, 2).text()
                existing_segment = self.tableWidget.cellWidget(r, 1)
                if existing_segment and isinstance(existing_segment, QComboBox):
                    existing_segment_value = existing_segment.currentText()
                    if (project_id == existing_project and
                        subtask == existing_subtask and
                        selected_segment == existing_segment_value):
                        combo_box.setCurrentIndex(0)
                        QMessageBox.warning(self,"Duplicate Entry","The combination already exists")
                        return



    def populate_table_by_week(self):
        user=self.label_3.text() 

        

        week = int(self.week_combo.currentText())
        
        # db=GetCursor()
        # if db.conn and db.cursor:
        #     query="select project,segment,subtask,weekday_1,weekday_2,weekday_3,weekday_4,weekday_5,weekday_6,weekday_7 from tasks where week=%s and username=%s"
        #     values=(week,user)
        #     db.cursor.execute(query,values)
        #     a=db.cursor.fetchall()
        #     self.populate_rows(a)
        # db.close_connection()
        # print("yess i am being hit")

        query=text("select project,segment,subtask,weekday_1,weekday_2,weekday_3,weekday_4,weekday_5,weekday_6,weekday_7 from tasks where week=:week and username=:user")
        params={"week":week,"user":user}
        worker=DatabaseWorker((query,params))
        worker.result_ready.connect(self.populate_rows)
        worker.error_occured.connect(self.handle_error)
        worker.finished.connect(lambda:(self.threads.remove(worker),self.cal_daily_weekly_totals()))
        self.threads.append(worker)
        worker.start()
    def populate_table_on_state(self):
        user=self.label_3.text() 
        # def get_week():
        #     query=text("select week from user_state where username=:user")
        #     params={"user":user}
        #     result={"week":1}
        #     worker=DatabaseWorker((query,params))
        #     loop=QEventLoop()
        #     def handle_result(data):
        #         if data:
        #             result["week"]=data[0][0]
        #             loop.quit()

        #     worker.result_ready.connect(handle_result)
        #     worker.error_occured.connect(self.handle_error)
        #     worker.finished.connect(lambda:self.threads.remove(worker))
        #     self.threads.append(worker)
        #     worker.start()
        #     loop.exec()
        #     return result["week"]
        
        def get_week_2():
            query=text("select week from user_state where username=:user")
            params={"user":user}
            result={"week":1}
            worker=DatabaseWorker((query,params))
            loop = QEventLoop()
            def handle_result(data):
                if data:
                    result["week"]=data[0][0]
                
            

            worker.result_ready.connect(handle_result)
            worker.error_occured.connect(self.handle_error)
            worker.finished.connect(lambda:(self.threads.remove(worker),loop.quit()))
            self.threads.append(worker)
            worker.start()
            loop.exec()
            return result["week"]







        week = get_week_2()
        self.week_combo.setCurrentText(str(week))
        print("the week is ",week)
      
        query=text("select project,segment,subtask,weekday_1,weekday_2,weekday_3,weekday_4,weekday_5,weekday_6,weekday_7 from tasks where week=:week and username=:user")
        params={"week":week,"user":user}
        worker=DatabaseWorker((query,params))
        worker.result_ready.connect(self.populate_rows)
        worker.error_occured.connect(self.handle_error)
        worker.finished.connect(lambda:(self.threads.remove(worker),self.cal_daily_weekly_totals()))
        self.threads.append(worker)
        worker.start()
        # db=GetCursor()
        # if db.conn and db.cursor:
        #     query="select project,segment,subtask,weekday_1,weekday_2,weekday_3,weekday_4,weekday_5,weekday_6,weekday_7 from tasks where week=%s and username=%s"
        #     values=(week,user)
        #     db.cursor.execute(query,values)
        #     a=db.cursor.fetchall()
        #     self.populate_rows(a)
        # db.close_connection()
        # print("yess i am being hit")
    


    def handle_error(self,error):
        QMessageBox.warning(self,"Error",f"{error}")
    def populate_rows(self,data):
        row=len(data)
        
        self.tableWidget.setRowCount(len(data))
        for row_idx,row_data in enumerate(data):
            for col_idx, val in enumerate(row_data):
                if val is not None and val != '':
                    if col_idx==1:
                        combo_box=QComboBox()
                        combo_box.addItems(["-","IH-PD","BU Doors","BU Windows","BU Facades","PM","TD","DHI","BU Sliders","CE-V8 Process","CE - CIP","Fabrication","Testing","SIPL"])
                        self.tableWidget.setCellWidget(row_idx,col_idx,combo_box)

                        combo_box.setCurrentText(str(val))
                        combo_box.setEnabled(False)
                        combo_box.currentIndexChanged.connect(self.handle_2)

                    else:
                        item=QTableWidgetItem(str(val))
                        self.tableWidget.setItem(row_idx,col_idx,item)

                    # print("item inserted")
                else:
                    item=QTableWidgetItem("")
                    self.tableWidget.setItem(row_idx,col_idx,item)
        

    def handle_2(self):
        combo_box = self.sender()
        if not isinstance(combo_box, QComboBox):
            return
        
        for row in range(self.tableWidget.rowCount()):
            if self.tableWidget.cellWidget(row, 1) == combo_box:
                current_row = row
                break
        else:
            return
        project_id_item = self.tableWidget.item(current_row, 0)
        subtask_item = self.tableWidget.item(current_row, 2)
        selected_segment = combo_box.currentText()
        if project_id_item is None or subtask_item is None:
            return  
        project_id = project_id_item.text()
        subtask = subtask_item.text()
        for r in range(self.tableWidget.rowCount()):
            if r != current_row:  # Skip the current row
                existing_project_item = self.tableWidget.item(r, 0)
                existing_subtask_item = self.tableWidget.item(r, 2)
                existing_segment_combo = self.tableWidget.cellWidget(r, 1)

                if (existing_project_item and existing_subtask_item and
                    isinstance(existing_segment_combo, QComboBox)):
                    existing_project = existing_project_item.text()
                    existing_subtask = existing_subtask_item.text()
                    existing_segment = existing_segment_combo.currentText()
                    if (project_id == existing_project and
                        subtask == existing_subtask and
                        selected_segment == existing_segment):
                        combo_box.setCurrentIndex(0)
                        QMessageBox.warning(self, "Duplicate Entry", "The combination already exists")
                        return

                        




    def add_to_table(self,project,subtask):
        row_count=self.tableWidget.rowCount()
        self.tableWidget.insertRow(row_count)
        
        self.tableWidget.setItem(row_count,0,QTableWidgetItem(project))
        self.tableWidget.setItem(row_count,1,QTableWidgetItem(subtask))



    def get_time(self,week_number):
        first_day_of_year = datetime(datetime.now().year+1, 1, 1)
        start_of_week = first_day_of_year + timedelta(weeks=week_number - 1)
        for i in range(7):  # Assuming 7 days per week
            day_date = start_of_week + timedelta(days=i)
            header = day_date.strftime("%b %d-%a")  # Format: "Dec 01-Mon"
            print(header,end=" ")
    def update_table_headers(self,week_number):
    # Get the first day of the year
        year = datetime.now().year+1
        first_day_of_year = datetime(year, 1, 1)

    # Find the first Monday before or on January 1
        if first_day_of_year.weekday() != 0:  # If not Monday
            first_monday = first_day_of_year - timedelta(days=first_day_of_year.weekday())
        else:
            first_monday = first_day_of_year

        # Calculate the start of the given week
        start_of_week = first_monday + timedelta(weeks=week_number - 1)


        for i in range(7):
            day_date = start_of_week + timedelta(days=i)
            header = day_date.strftime("%b %d-%a")  # Format: "Jan 01-Mon"
            self.tableWidget.setHorizontalHeaderItem(3+i,QTableWidgetItem(header))
            month_date=header.split('-')[0]
            label=getattr(self,f"day{i+1}")
            label.setText(month_date)
    
    def update_table_headers_2(self, week_number):
        year = datetime.now().year
        first_day_of_year = datetime(year, 1, 1)
        if first_day_of_year.weekday() != 0:
            first_monday = first_day_of_year - timedelta(days=first_day_of_year.weekday())
        else:
            first_monday = first_day_of_year
        start_of_week = first_monday + timedelta(weeks=week_number - 1)
        days_in_year = (datetime(year + 1, 1, 1) - datetime(year, 1, 1)).days  # 365 or 366 days
        end_of_year = datetime(year, 12, 31)
        for i in range(7):
            day_date = start_of_week + timedelta(days=i)
            if day_date > end_of_year:
                break
            header = day_date.strftime("%b %d-%a")
            self.tableWidget.setHorizontalHeaderItem(3 + i, QTableWidgetItem(header))
            month_date = header.split('-')[0]
            label = getattr(self, f"day{i + 1}")
            label.setText(month_date)

        if week_number == 53 and start_of_week > end_of_year:
            for i in range(7):
                day_date = start_of_week + timedelta(days=i)
                if day_date > end_of_year:
                    break
                header = day_date.strftime("%b %d-%a")
                self.tableWidget.setHorizontalHeaderItem(2 + i, QTableWidgetItem(header))
                month_date = header.split('-')[0]
                label = getattr(self, f"day{i + 1}")
                label.setText(month_date)


    def update_projects(self):
        user=self.label_3.text()
        print("the user is ",user)
        # db=GetCursor()
        # if db.conn and db.cursor:
        #     query="select distinct project from assignments where username=%s"
        #     db.cursor.execute(query,(user,))
        #     a=db.cursor.fetchall()
        #     projects=[p[0]for p in a]
        #     self.project_combo.addItems(projects)
        #     self.project_combo.addItem("Non-Billable")
        # db.close_connection()
        def populate_projects(obj):
            projects=[p[0] for p in obj]
            self.project_combo.clear()
            self.project_combo.addItems(projects)
            self.project_combo.addItem("Non-Billable")

        query=text("select distinct project from assignments where username=:user")
        params={"user":user}
        worker=DatabaseWorker((query,params))
        worker.result_ready.connect(populate_projects)
        worker.error_occured.connect(self.handle_error)
        worker.finished.connect(lambda:self.threads.remove(worker))
        self.threads.append(worker)
        worker.start()
    def update_subtasks(self):
        user=self.label_3.text()
        project=self.project_combo.currentText()
        if project and project!="Non-Billable":
            # db=GetCursor()
            # if db.conn and db.cursor:
            #     query="select  subtask from assignments where project=%s and username=%s"
            #     db.cursor.execute(query,(project,user))
            #     a=db.cursor.fetchall()
            #     subtasks=[p[0]for p in a]
            #     print(subtasks)
            #     self.subtask_combo.clear()
            #     self.subtask_combo.addItems(subtasks)
            #     db.close_connection()
            def populate_subtasks(obj):
                subtasks=[subtask[0] for subtask in obj]
                self.subtask_combo.clear()
                self.subtask_combo.addItems(subtasks)

            query=text("select  subtask from assignments where project=:project and username=:user")
            params={"project":project,"user":user}
            worker=DatabaseWorker((query,params))
            worker.result_ready.connect(populate_subtasks)
            worker.error_occured.connect(self.handle_error)
            worker.finished.connect(lambda:self.threads.remove(worker))
            self.threads.append(worker)
            worker.start()


        else:
            items=["Miscellaneous","Training","Annual Leave","Sick/Casual Leave","Declared Holiday","Restricted Holiday"]
            self.subtask_combo.clear()
            self.subtask_combo.addItems(items)

            


        
        
def load_stylesheet(file_path):
    """Load the QSS stylesheet from the given file path."""
    try:
        with open(file_path, "r") as file:
            return file.read()
    except Exception as e:
        print(f"Error loading stylesheet: {e}")
        return ""
# app=QApplication(sys.argv)
# stylesheet=load_stylesheet("styles.qss")
# app.setStyleSheet(stylesheet)
# window=MyEmployee()
# window.show()
# app.exec()